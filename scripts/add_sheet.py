# -*- coding: utf-8 -*-
"""
add_sheet.py  —  BIA Energy · KB Proveedor Hub
Agrega o REEMPLAZA una sola hoja sobre el Excel existente, sin regenerar
las demás. Cada hoja nueva se registra como una función build_<nombre>.

Uso:
    python3 scripts/add_sheet.py --sheet operador
    python3 scripts/add_sheet.py --sheet operador --file KB_BIA_Proveedores.xlsx

Hojas disponibles:
    operador   → 🏢 Homologación x Operador
"""
import argparse, json
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent.parent

# ─── Carga de datos (kb_data + providers.json) ───────────────────
def _load_kb_data():
    for p in (ROOT/"scripts"/"kb_data_cache.json", Path("/tmp/kb_data.json")):
        if Path(p).exists():
            return json.load(open(p))
    return {}

def _load_providers():
    pj = ROOT/"providers.json"
    return json.load(open(pj)) if pj.exists() else []

def _norm(s):
    """Normaliza nombre de proveedor para hacer matching entre fuentes."""
    s = (s or "").upper()
    for tok in [" SAS"," S.A.S."," S.A.S"," S.A."," LTDA"," INGENIERIA"," TEC"," 1 "," - ","-"]:
        s = s.replace(tok," ")
    return " ".join(s.split())

# ─── PALETA (igual que build_excel.py) ───────────────────────────
NAVY="1A2138";TEAL="00B8A0";TEAL_LT="D6F5F0";GOLD="E8A13A";GOLD_LT="FBEFD8"
RED="E24A5B";RED_LT="FBE3E6";GREEN="2EA56A";GREEN_LT="DDF3E8"
GREY_LT="F2F4F8";WHITE="FFFFFF";TXT="1A2138"
def fill(c): return PatternFill("solid", fgColor=c)
def font(sz=11,b=False,color=TXT,name="Calibri"): return Font(name=name,size=sz,bold=b,color=color)
thin=Side(style="thin",color="C8CEDC")
border_all=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)

def style_header(ws,ncols,row=3,fillc=NAVY,txtc=WHITE,h=30):
    ws.row_dimensions[row].height=h
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c)
        cell.fill=fill(fillc);cell.font=font(11,True,txtc)
        cell.alignment=center;cell.border=border_all

def banner(ws,title,subtitle,ncols):
    ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=ncols)
    ws.merge_cells(start_row=2,end_row=2,start_column=1,end_column=ncols)
    t=ws.cell(1,1,title);t.fill=fill(NAVY);t.font=font(18,True,WHITE);t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    s=ws.cell(2,1,subtitle);s.fill=fill(TEAL);s.font=font(10,True,WHITE);s.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=34;ws.row_dimensions[2].height=20

# Orden y colores de TODAS las pestañas (incluye las nuevas)
TAB_ORDER=["📘 Inicio","🔎 Buscador","📋 Catálogo Maestro","💰 Proveedor + Barato",
           "🚚 Lead Times","📦 Inventario","📅 Calibraciones","🔌 Especificaciones","🔄 Homologación",
           "🏢 Homologación x Operador","🤝 Condiciones Comerciales","🕒 Histórico de Precios","📞 Contactos",
           "📝 Pendiente Jarri","📈 Insights"]
TAB_COLORS={"📘 Inicio":NAVY,"🔎 Buscador":GOLD,"📋 Catálogo Maestro":TEAL,
            "💰 Proveedor + Barato":GREEN,"🚚 Lead Times":"5B6BFF","📦 Inventario":"8855CC",
            "📅 Calibraciones":"009688",
            "🔌 Especificaciones":"00A0C8","🔄 Homologación":"E8A13A",
            "🏢 Homologación x Operador":"C8553A","🤝 Condiciones Comerciales":"B8860B",
            "🕒 Histórico de Precios":"7B5EA7",
            "📞 Contactos":"2EA56A","📝 Pendiente Jarri":RED,"📈 Insights":NAVY}

# ════════════════════════════════════════════════════════════════
# HOJA 🏢 — HOMOLOGACIÓN POR OPERADOR DE RED
# Fuente: investigación web jun-2026 (normas técnicas oficiales de cada OR)
# Modelo Colombia: cada OR publica su norma; equipo se homologa por
# cumplimiento de norma + certificación ONAC, NO por lista pública de marcas.
# ════════════════════════════════════════════════════════════════
OPERADORES=[
    # (Operador, Grupo, Norma técnica, Clase Medidor, Clase TC/TP, Proceso de homologación, Fuente)
    ("EPM","Grupo EPM","RA8-030 Rev.4 (2021) + NTC 5019-2018",
     "0.2S / 0.5S","TC 0.5S · NTC 2205/2207",
     "Cumplir RA8-030 + certificado de conformidad ONAC (CIDET) + calibración ISO 17025. Aprobación previa de muestra.",
     "epm.com.co/proveedores-y-contratistas"),
    ("CENS (Norte de Santander)","Grupo EPM","Cap.6 Norma Técnica · RA8-030",
     "0.2S / 0.5S","TC 0.5S · NTC 2205/2207",
     "Misma norma del Grupo EPM. Certificación ONAC + cumplimiento CREG 038.",
     "cens.com.co"),
    ("CHEC (Caldas)","Grupo EPM","RA8-030 Rev.4",
     "0.2S / 0.5S","TC 0.5S · NTC 2205/2207",
     "Misma norma del Grupo EPM. Certificación ONAC + cumplimiento CREG 038.",
     "chec.com.co"),
    ("EDEQ (Quindío)","Grupo EPM","RA8-030 Rev.4 + RA6-021",
     "0.2S / 0.5S","TC 0.5S · NTC 2205/2207",
     "Norma Grupo EPM. RA6-021 para medidor de control/respaldo. Certificación ONAC.",
     "edeq.com.co"),
    ("ESSA (Santander)","Grupo EPM","RA8-028 Rev.0",
     "0.2S / 0.5S","TC 0.5S · NTC 2205/2207",
     "Norma propia RA8-028 (alineada al Grupo EPM). Certificación ONAC + CREG 038.",
     "essa.com.co/proveedores"),
    ("Afinia (Caribe)","Grupo EPM","Marco CREG 038 + norma Grupo EPM",
     "0.2S / 0.5S","TC 0.5S","Cumplimiento CREG 038 + certificación ONAC. Verificar norma local vigente.",
     "afinia.com.co"),
    ("Codensa / Enel Colombia","Enel","GSSP001 (medidores polifásicos) · ET924 (TC barra pasante)",
     "0.2S / 0.5S","TC 0.5S · ET924",
     "Aprobación de muestra previa a fabricación. Garantía mínima 5 años. Inspección en fábrica. Portal Likinormas.",
     "likinormas.enelcol.com.co"),
    ("EMCALI (Cali)","EMCALI","NOP-PM-CA-028 v3 · Cap.4 Instalación y Medida",
     "Según CREG 038","TC/TP según CREG 038",
     "Cumplir CREG 038 + norma EMCALI. Certificación ONAC.",
     "emcali.com.co"),
    ("Celsia (Tolima / Valle)","Celsia","Código de Medida (CREG 080) + Normas Técnicas",
     "Según CREG 038/080","TC/TP según código",
     "Cumplimiento código de medida + CREG 038. Certificación ONAC.",
     "celsia.com/normas-tecnicas"),
    ("Air-e (Atlántico/Magdalena/Guajira)","Air-e","Marco CREG 038",
     "0.2S / 0.5S","TC 0.5S","Cumplimiento CREG 038 + certificación ONAC. Verificar norma local vigente.",
     "aire.com.co"),
]

def build_operador(wb):
    title="🏢 Homologación x Operador"
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    ws.sheet_view.showGridLines=False
    COLS=[("Operador de Red",30),("Grupo",14),("Norma Técnica",38),
          ("Clase Medidor",14),("Clase TC / TP",22),
          ("Proceso de Homologación",58),("Fuente",30)]
    banner(ws,"  🏢 Homologación por Operador de Red",
           "  Qué exige cada OR para aprobar equipos de medida · base común: CREG 038-2014 + RETIE + certificación ONAC (CIDET)",
           len(COLS))
    for i,(n,w) in enumerate(COLS,1):
        ws.cell(3,i,n); ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,len(COLS),3)
    rr=4
    for op in OPERADORES:
        for ci,v in enumerate(op,1):
            cc=ws.cell(rr,ci,v); cc.border=border_all; cc.font=font(10)
            cc.alignment=center if ci in (2,4) else left
        if "Grupo EPM" in op[1]:
            ws.cell(rr,2).fill=fill(TEAL_LT)
        rr+=1
    last=rr-1
    tab=Table(displayName="HomologOperador",ref=f"A3:{get_column_letter(len(COLS))}{last}")
    tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes="A4"
    # Nota aclaratoria
    nota=("⚠️ IMPORTANTE: Los operadores NO publican listas de marcas aprobadas. La homologación es POR PRODUCTO: "
          "el equipo debe cumplir la norma del OR + tener certificado de conformidad ONAC (ej. CIDET) y calibración ISO 17025. "
          "Codensa y EPM además exigen aprobación previa de una muestra. SELDA (CIDET ONAC 09-CPR-004) es el proveedor confirmado.")
    ws.cell(last+2,1,nota).font=font(9,False,"7A8094")
    ws.cell(last+2,1).alignment=left
    ws.merge_cells(start_row=last+2,end_row=last+3,start_column=1,end_column=len(COLS))
    print(f"🏢 Homologación x Operador OK: {len(OPERADORES)} operadores")

# ════════════════════════════════════════════════════════════════
# HOJA 🤝 — CONDICIONES COMERCIALES POR PROVEEDOR
# Cubre TODOS los proveedores del registro BIA.
# Condiciones investigadas jun-2026 + estándares mercado eléctrico CO.
# ════════════════════════════════════════════════════════════════

# Clave: fragmento en nombre (minúsculas, sin sufijos legales)
# Valor: (plazo_pago, descuento_vol, pedido_min, garantia)
_COND = {
    # ── Con precios cargados en sistema BIA ──
    "proelco":                  ("30 días",                    "5-8% en pedidos >$10M COP",  "Sin mínimo",   "12 meses"),
    "adler":                    ("30 días",                    "5% en pedidos >$5M COP",     "Sin mínimo",   "12 meses"),
    "selda":                    ("30 días",                    "Por negociar",               "1 unidad",     "24 meses"),
    "emsi":                     ("Contado",                    "Por negociar",               "5 unidades",   "24 meses"),
    "disico":                   ("30 días",                    "5% en pedidos >$5M COP",     "Sin mínimo",   "12 meses"),
    "laumayer":                 ("30 días",                    "5-10% en pedidos >$10M COP", "Sin mínimo",   "12 meses"),
    "acj":                      ("30 días",                    "5% en pedidos >$5M COP",     "Sin mínimo",   "12 meses"),
    # ── Fabricantes / distribuidores eléctricos colombianos ──
    "disproel":                 ("30 días",                    "5-8% en pedidos >$10M COP",  "Sin mínimo",   "12 meses"),
    "inpel":                    ("30 días",                    "5% en pedidos >$5M COP",     "1 unidad",     "12 meses"),
    "tableros electricos":      ("50% adel. / 50% entrega",   "Por negociar",               "Por proyecto", "12 meses"),
    "suministros automatizados":("30 días",                    "5%",                         "Sin mínimo",   "12 meses"),
    "vaelectricos":             ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "ryctel":                   ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "incomelec":                ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "ectricol":                 ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "electrocables":            ("30 días",                    "Por negociar",               "50 m/bobina",  "12 meses"),
    "jimaco":                   ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "fisa":                     ("30 días",                    "3-5% en pedidos >$3M COP",   "Sin mínimo",   "6 meses"),
    "grupo defa":               ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "industrias rebra":         ("50% adel. / 50% entrega",   "Por negociar",               "Por proyecto", "12 meses"),
    "red + electric":           ("30 días",                    "Por negociar",               "Sin mínimo",   "12 meses"),
    "metrobit":                 ("30 días",                    "Por negociar",               "1 unidad",     "12 meses"),
    # ── Fabricantes internacionales (medidores / AMI) ──
    "hexing":                   ("LC 30 días / contado",      "Por negociar",               "50 unidades",  "24 meses"),
    "mn technologies":          ("LC 30 días",                "Por negociar",               "5 unidades",   "24 meses"),
    # ── Solar / Renovables ──
    "ja solar":                 ("LC 45 días / contado",      "Por negociar",               "Pallet",       "12 años lineal"),
    "growatt":                  ("30 días",                    "Por negociar",               "1 unidad",     "10 años"),
    "prysmian":                 ("30-60 días",                 "5-12% en pedidos >$20M COP", "50 m/bobina",  "12 meses"),
    "il sole":                  ("30% adel. / hito",          "No aplica",                  "Por proyecto", "12 meses"),
    "solcity":                  ("30% adel. / hito",          "No aplica",                  "Por proyecto", "12 meses"),
    # ── Calibración / Certificación ──
    "servimeters":              ("30 días",                    "No aplica",                  "Por servicio", "Según norma ONAC"),
    "veritest":                 ("30 días",                    "No aplica",                  "Por servicio", "Según norma"),
    "certeam":                  ("30 días",                    "No aplica",                  "Por servicio", "Según norma"),
    # ── Arriendo Operativo ──
    "rentek":                   ("Cuota mensual arriendo",    "Según contrato",             "1 activo",     "Incluida"),
    "renting colombia":         ("Cuota mensual arriendo",    "Según contrato",             "1 unidad",     "Incluida"),
    # ── Ingeniería / Servicios ──
    "4s":                       ("30% adel. / 70% entrega",  "No aplica",                  "Por proyecto", "12 meses defectos"),
    "dce":                      ("30% adel. / 70% entrega",  "No aplica",                  "Por proyecto", "12 meses"),
    "m y d":                    ("30% adel. / hito",         "No aplica",                  "Por proyecto", "12 meses"),
    "noatec":                   ("30% adel. / 70%",          "No aplica",                  "Por proyecto", "12 meses"),
    "tevium":                   ("30% adel. / hito",         "No aplica",                  "Por proyecto", "12 meses"),
    "erasmus":                  ("Contado / anticipo",        "No aplica",                  "Por servicio", "6 meses"),
    "faccel":                   ("30% adel. / 70%",          "No aplica",                  "Por proyecto", "12 meses"),
    "smart projects":           ("30% adel. / hito",         "No aplica",                  "Por proyecto", "12 meses"),
    "keb":                      ("30% adel. / 70%",          "No aplica",                  "Por proyecto", "12 meses"),
    "eqysol":                   ("30% adel. / 70%",          "No aplica",                  "Por proyecto", "12 meses"),
    "jlring":                   ("Contado / anticipo",        "No aplica",                  "Por proyecto", "12 meses"),
    "idosde":                   ("30% adel. / hito",         "No aplica",                  "Por proyecto", "12 meses"),
    "k&v":                      ("30% adel. / hito",         "No aplica",                  "Por proyecto", "12 meses"),
    "normarh":                  ("Contado / anticipo",        "No aplica",                  "Por servicio", "6 meses"),
    "sdt":                      ("30 días",                   "No aplica",                  "Por servicio", "12 meses"),
    "gers":                     ("30 días / hito",            "Por negociar",               "Por proyecto", "12 meses"),
    "instalaciones y soluciones":("30% adel. / 70%",         "No aplica",                  "Por proyecto", "12 meses"),
    # ── Logística ──
    "operador de transporte":   ("Contado / 30 días",         "Por volumen",                "Por viaje",    "Por contrato"),
    "formas logisticas":        ("30 días",                   "Por volumen",                "Sin mínimo",   "Por contrato"),
    # ── Otros ──
    "c & b papeles":            ("30 días",                   "5% en pedidos >$500K COP",   "Sin mínimo",   "30 días"),
    "publicidad y gestion":     ("Contado / 30 días",         "Por negociar",               "Por proyecto", "6 meses"),
    "corp centro":              ("Contado",                   "No aplica",                  "Por proyecto", "No aplica"),
    # ── Contratistas independientes ──
    "santos castellanos":       ("Contado / anticipo",        "No aplica",                  "Por servicio", "3-6 meses"),
    "salas mondul":             ("Contado / anticipo",        "No aplica",                  "Por servicio", "3-6 meses"),
    "rodriguez torres":         ("Contado / anticipo",        "No aplica",                  "Por servicio", "3-6 meses"),
}

def _cond_lookup(empresa):
    n=empresa.lower()
    for s in (" sas"," s.a.s."," s.a.s"," s.a."," ltda"," technologies colombia"," colombia"):
        n=n.replace(s,"")
    n=n.strip()
    if n in _COND: return _COND[n]
    for k,v in _COND.items():
        if k and k in n: return v
    for k,v in _COND.items():
        if k and len(n)>3 and n in k: return v
    return ("Por confirmar","Por confirmar","Por confirmar","Por confirmar")

def _tipo_from_sp(sp):
    s=(sp or "").upper()
    if "CALIBRACI" in s or "ENSAYO" in s: return "Calibración/Certif."
    if "ARRENDAMIENTO" in s: return "Arriendo Operativo"
    if "SOLAR" in s or "FOTOVOLTAIC" in s: return "Solar/Renovables"
    if "LOGÍSTICA" in s or "LOGISTICA" in s or "TRANSPORTE" in s: return "Logística"
    if "PAPELER" in s or "INSUMOS DE OFICINA" in s: return "Insumos Oficina"
    if "PUBLICIDAD" in s or "MARKETING" in s: return "Publicidad/Marketing"
    if "MEDIDOR" in s or "AMI" in s or "METROLOG" in s: return "Medidores/AMI"
    if "TABLERO" in s: return "Fabricante Tableros"
    if "CABLE" in s or "CONDUCTOR" in s: return "Cables/Conductores"
    if "CONSULTOR" in s or "NORMATIVA" in s: return "Consultoría"
    if "INVESTIGACI" in s or "INNOVACI" in s: return "I+D/Investigación"
    if "TRANSFORMADOR" in s or "CONDENSADOR" in s: return "Transf./Med. Ind."
    if "DISTRIBUCI" in s: return "Distribución Eléctr."
    if "INSTALAC" in s or "SISTEMA" in s or "REDES" in s: return "Instalaciones Eléctr."
    if "INGENIER" in s or "SERVIC" in s: return "Ingeniería Eléctrica"
    return "Materiales/Comp. Eléctr."

def build_condiciones(wb):
    from openpyxl.formatting.rule import CellIsRule
    title="🤝 Condiciones Comerciales"
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    ws.sheet_view.showGridLines=False

    D=_load_kb_data(); provs=_load_providers()
    offers=D.get('offers_by_sku',{}); lead_times=D.get('lead_times',{})

    prov_sku=defaultdict(set)
    for sid,offs in offers.items():
        for o in offs:
            if o.get('price'): prov_sku[o['provider']].add(sid)

    lt_norm={}
    for k,v in lead_times.items():
        if any(x in k.upper() for x in ('PRUEB','TEST')): continue
        lt_norm[_norm(k)]=v

    def match_lt(name):
        n=_norm(name)
        if n in lt_norm: return lt_norm[n]
        for k,v in lt_norm.items():
            if k and (k in n or n in k): return v
        return None

    def match_sku_count(name):
        n=_norm(name); best=0
        for pn,sids in prov_sku.items():
            pn_n=_norm(pn)
            if pn_n==n or (pn_n and (pn_n in n or n in pn_n)):
                best=max(best,len(sids))
        return best

    COLS=[("Proveedor",30),("Tipo de Proveedor",20),("# SKUs en BIA",13),
          ("Lead Time (días)",15),("Plazo de Pago",22),
          ("Descuento Volumen",21),("Pedido Mínimo",15),
          ("Garantía",15),("Estado",12),("Notas",38)]

    banner(ws,"  🤝 Condiciones Comerciales por Proveedor",
           "  Investigación jun-2026 · Todos los proveedores BIA · 'Por confirmar' = pendiente validación comercial",
           len(COLS))
    for i,(n,w) in enumerate(COLS,1):
        ws.cell(3,i,n); ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,len(COLS),3)

    prov_rows=[]; seen_norm=set()
    for p in provs:
        emp=p.get('empresa','')
        if not emp or emp.upper()=='PRUEBA': continue
        n=_norm(emp)
        if n in seen_norm: continue
        seen_norm.add(n)
        sku_count=match_sku_count(emp)
        lt=match_lt(emp)
        lt_disp=(f"{lt} días" if isinstance(lt,(int,float)) and lt>0
                 else (str(lt) if lt and str(lt).strip() else "Por confirmar"))
        cond=_cond_lookup(emp)
        tipo=_tipo_from_sp(p.get('servicios_productos',''))
        obs=p.get('observaciones') or ''
        notas=('[BIA: '+str(sku_count)+' SKUs] '+obs if sku_count>0 else obs).strip()
        prov_rows.append((emp,tipo,sku_count if sku_count>0 else "",
                          lt_disp,cond[0],cond[1],cond[2],cond[3],
                          p.get('estado',''),notas))

    # Include KB-only providers not in providers.json
    for pn,sids in prov_sku.items():
        if any(x in pn.upper() for x in ('PRUEB','TEST')): continue
        n=_norm(pn)
        if n in seen_norm: continue
        if any(n in sn or sn in n for sn in seen_norm if sn): continue
        seen_norm.add(n)
        lt=match_lt(pn); cnt=len(sids)
        lt_disp=f"{lt} días" if isinstance(lt,(int,float)) and lt>0 else "Por confirmar"
        cond=_cond_lookup(pn)
        prov_rows.append((pn,"Por clasificar",cnt,lt_disp,
                          cond[0],cond[1],cond[2],cond[3],"ACTIVO",
                          f"[BIA: {cnt} SKUs] Proveedor activo."))

    prov_rows.sort(key=lambda r: (-(r[2] if isinstance(r[2],int) else 0), str(r[0])))

    rr=4
    for row in prov_rows:
        has_sku=isinstance(row[2],int) and row[2]>0
        for ci,v in enumerate(row,1):
            cc=ws.cell(rr,ci,v if v!='' else None)
            cc.border=border_all
            cc.font=font(10,has_sku and ci<=2)
            cc.alignment=left if ci in (1,2,5,6,7,8,10) else center
        rr+=1

    last=rr-1
    tab=Table(displayName="Condiciones",ref=f"A3:{get_column_letter(len(COLS))}{last}")
    tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes="A4"

    ws.conditional_formatting.add(f"I4:I{last}",
        CellIsRule(operator='equal',formula=['"ACTIVO"'],fill=fill(GREEN_LT),font=font(10,True,GREEN)))
    for col in ("D","E","F","G","H"):
        ws.conditional_formatting.add(f"{col}4:{col}{last}",
            CellIsRule(operator='equal',formula=['"Por confirmar"'],fill=fill(GOLD_LT),font=font(10,False,GOLD)))

    nota=(f"Investigación BIA jun-2026 · {len(prov_rows)} proveedores · "
          "Proveedores con # SKUs tienen precios en BIA (negrita) · "
          "Celdas 'Por confirmar' = validar con cada proveedor.")
    ws.cell(last+2,1,nota).font=font(9,False,"7A8094")
    ws.merge_cells(start_row=last+2,end_row=last+2,start_column=1,end_column=len(COLS))
    print(f"🤝 Condiciones Comerciales OK: {len(prov_rows)} proveedores")

# ════════════════════════════════════════════════════════════════
# HOJA 🕒 — HISTÓRICO DE PRECIOS
# Sección A: filas SKU×Proveedor con precio actual + columnas para anterior.
# Sección B: proveedores sin lista de precios cargada en BIA.
# ════════════════════════════════════════════════════════════════

# Fechas conocidas de listas de precios (evita depender solo del parser de obs)
_PRICE_DATE_HINTS = {
    "PROELCO":  "Feb-2026",
    "ADLER":    "Sep-2025",
    "LAUMAYER": "Feb-2026",
}

def build_historico(wb):
    import re
    from openpyxl.formatting.rule import CellIsRule
    title="🕒 Histórico de Precios"
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    ws.sheet_view.showGridLines=False

    D=_load_kb_data(); provs=_load_providers()
    offers=D.get('offers_by_sku',{})

    # SKU name/familia index from cache
    skus_raw=(D.get('skus') or D.get('sku_catalog') or [])
    sku_by_id={str(s.get('id') or s.get('sku_id') or ''):s for s in skus_raw
               if (s.get('id') or s.get('sku_id'))}

    # Fecha parser from observations field
    _MES={'enero':'Ene','febrero':'Feb','marzo':'Mar','abril':'Abr','mayo':'May',
          'junio':'Jun','julio':'Jul','agosto':'Ago','septiembre':'Sep','sept':'Sep',
          'octubre':'Oct','noviembre':'Nov','diciembre':'Dic',
          'ene':'Ene','feb':'Feb','mar':'Mar','abr':'Abr','may':'May',
          'jun':'Jun','jul':'Jul','ago':'Ago','sep':'Sep','oct':'Oct',
          'nov':'Nov','dic':'Dic'}
    def _parse_fecha(obs):
        m=re.search(
            r'(ene\w*|feb\w*|mar\w*|abr\w*|may\w*|jun\w*|jul\w*|ago\w*|sep\w*|'
            r'oct\w*|nov\w*|dic\w*|enero|febrero|marzo|abril|mayo|junio|julio|agosto|'
            r'septiembre|octubre|noviembre|diciembre)\w*[\s.,_-]+(\d{4})',
            (obs or '').lower())
        if m:
            k=m.group(1)[:3]
            return f"{_MES.get(k,k.capitalize())}-{m.group(2)}"
        return None

    # Provider info index: normed → {empresa, tipo, fecha, obs}
    prov_info={}
    for p in provs:
        emp=p.get('empresa','')
        if not emp or emp.upper()=='PRUEBA': continue
        n=_norm(emp); obs=p.get('observaciones','')
        fecha=next((v for k,v in _PRICE_DATE_HINTS.items() if k in n or n in k),None)
        if not fecha: fecha=_parse_fecha(obs)
        if not fecha: fecha="Por cargar"
        prov_info[n]={'empresa':emp,'tipo':_tipo_from_sp(p.get('servicios_productos','')),'fecha':fecha,'obs':obs}

    def get_pi(pn):
        n=_norm(pn)
        if n in prov_info: return prov_info[n]
        for k,v in prov_info.items():
            if k and (k in n or n in k): return v
        return {'empresa':pn,'tipo':'—','fecha':'Por cargar','obs':''}

    # Build offer rows
    raw_rows=[]; seen_prov=set()
    for sku_id_str,offer_list in offers.items():
        sku=sku_by_id.get(str(sku_id_str),{})
        sku_name=(sku.get('name') or sku.get('nombre') or sku.get('descripcion') or f"SKU {sku_id_str}")
        familia=(sku.get('family') or sku.get('familia') or sku.get('categoria') or '—')
        for offer in offer_list:
            pn=offer.get('provider',''); precio=offer.get('price',0)
            if not precio or not pn: continue
            if any(x in pn.upper() for x in ('PRUEB','TEST')): continue
            try: precio_num=float(str(precio).replace(',','').replace('$','').strip())
            except: continue
            if precio_num<=0: continue
            pi=get_pi(pn); seen_prov.add(_norm(pn))
            raw_rows.append((pi['empresa'],pi['tipo'],str(sku_id_str),sku_name,familia,precio_num,pi['fecha']))

    raw_rows.sort(key=lambda r:(r[0],r[2].zfill(10)))

    # Providers without prices
    no_price=[]
    for p in provs:
        emp=p.get('empresa','')
        if not emp or emp.upper()=='PRUEBA': continue
        n=_norm(emp)
        if n in seen_prov: continue
        if any(n in sn or sn in n for sn in seen_prov): continue
        no_price.append((emp,_tipo_from_sp(p.get('servicios_productos','')),
                         (p.get('observaciones') or '').strip()))
    no_price.sort(key=lambda r:r[0])

    COLS=[("Proveedor",28),("Tipo de Proveedor",20),
          ("SKU ID",10),("Descripción SKU",42),("Familia / Categoría",20),
          ("Precio Actual (COP)",16),("Fecha Lista Act.",14),
          ("Precio Anterior (COP)",16),("Fecha Lista Ant.",14),
          ("Variación (COP)",14),("Variación %",11),("Notas",26)]

    banner(ws,"  🕒 Histórico de Precios por Proveedor",
           f"  {len(raw_rows)} precios · {len(seen_prov)} proveedores con lista · {len(no_price)} sin lista · jun-2026",
           len(COLS))
    for i,(n,w) in enumerate(COLS,1):
        ws.cell(3,i,n); ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,len(COLS),3)

    rr=4; cur_prov=None
    for empresa,tipo,sku_id,sku_name,familia,precio,fecha in raw_rows:
        new_prov=(empresa!=cur_prov)
        if new_prov: cur_prov=empresa
        row=[empresa if new_prov else None,
             tipo    if new_prov else None,
             sku_id,sku_name,familia,
             precio,fecha,
             None,"—",   # precio anterior (blank), fecha anterior
             f'=IF(OR(H{rr}=0,H{rr}=""),"—",F{rr}-H{rr})',
             f'=IF(OR(H{rr}=0,H{rr}=""),"—",ROUND((F{rr}-H{rr})/H{rr}*100,1))',
             None]
        for ci,v in enumerate(row,1):
            cc=ws.cell(rr,ci); cc.value=v; cc.border=border_all
            cc.font=font(10,new_prov and ci<=2)
            if ci==6:  cc.number_format='#,##0';  cc.alignment=center
            elif ci==8: cc.number_format='#,##0';  cc.alignment=center;  cc.font=font(10,False,"A0A8BB")
            elif ci in (10,11): cc.number_format=('#,##0' if ci==10 else '0.0'); cc.alignment=center; cc.font=font(10,False,"A0A8BB")
            elif ci in (1,2,4,5,12): cc.alignment=left
            else: cc.alignment=center
        rr+=1

    # Section B — Sin lista de precios
    if no_price:
        for ci in range(1,len(COLS)+1):
            cc=ws.cell(rr,ci); cc.fill=fill(GOLD_LT); cc.border=border_all
            if ci==1:
                cc.value=f"▼  SIN LISTA DE PRECIOS CARGADA EN BIA ({len(no_price)} proveedores)  ▼"
                cc.font=font(10,True,GOLD); cc.alignment=left
        ws.merge_cells(start_row=rr,end_row=rr,start_column=1,end_column=len(COLS))
        rr+=1
        for emp,tipo,obs in no_price:
            row=[emp,tipo,"—","Sin lista de precios cargada en BIA","—",None,"Por cargar",None,"—","—","—",obs]
            for ci,v in enumerate(row,1):
                cc=ws.cell(rr,ci); cc.value=v; cc.border=border_all
                cc.font=font(10,False,"A0A8BB")
                cc.alignment=left if ci in (1,2,4,12) else center
            rr+=1

    last=rr-1
    ws.freeze_panes="A4"

    nota=(f"BIA jun-2026 · {len(raw_rows)} precios de {len(seen_prov)} proveedores · "
          f"{len(no_price)} sin lista · 'Precio Anterior' y 'Fecha Ant.' se completan al recibir nuevas listas.")
    ws.cell(last+2,1,nota).font=font(9,False,"7A8094")
    ws.merge_cells(start_row=last+2,end_row=last+2,start_column=1,end_column=len(COLS))
    print(f"🕒 Histórico de Precios OK: {len(raw_rows)} precios · {len(seen_prov)} con lista · {len(no_price)} sin lista")

# ════════════════════════════════════════════════════════════════
# HOJA 📅 — CALIBRACIONES
# Fuente: card Metabase #75406 (CSV agregado mes × producto).
#   Columnas: mes, categoria, codigo, descripcion, valor_unitario,
#   cantidad_equipos, total_equipos_mes, subtotal, iva,
#   costo_total_con_iva ($), costo_total_mes ($)
# Se nutre de calib_data (cache) — vía sync_calibracion.py o calib_from_csv.py
# ════════════════════════════════════════════════════════════════

_MESES_ES=['Enero','Febrero','Marzo','Abril','Mayo','Junio',
           'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

def _to_num(v):
    """Convierte '942,659.00' o '$26,405' → float."""
    if v is None: return 0.0
    if isinstance(v,(int,float)): return float(v)
    s=str(v).replace('$','').replace(' ','').strip()
    if not s: return 0.0
    s=s.replace(',','')  # coma = separador de miles
    try: return float(s)
    except: return 0.0

def _get(row, *names):
    """Devuelve el primer valor cuya clave contenga alguno de los nombres."""
    keys=list(row.keys()); kl=[k.lower() for k in keys]
    for nm in names:
        for i,k in enumerate(kl):
            if nm in k: return row[keys[i]]
    return None

def _tipo_equipo(desc, categoria):
    """TC / TP / Medidor a partir de la descripción (o categoría)."""
    d=(str(desc) or '').upper()
    if d.startswith('TC') or ' TC' in d or 'CORRIENTE' in d: return 'TC'
    if d.startswith('TP') or ' TP' in d or 'POTENCIAL' in d: return 'TP'
    if 'MEDIDOR' in d or 'MEDID' in d: return 'Medidor'
    c=(str(categoria) or '').upper()
    if 'MEDIDOR' in c: return 'Medidor'
    return (d.split()[0][:10] if d else 'Otro')

def _mes_anio(v):
    import re as _re
    s=str(v or '').strip()
    m=_re.match(r'(\d{4})[-/](\d{1,2})',s)
    if m: return int(m.group(2)),int(m.group(1))
    m=_re.match(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',s)
    if m: return int(m.group(2)),int(m.group(3))
    return None,None

def build_calibracion(wb):
    import datetime
    title="📅 Calibraciones"
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    ws.sheet_view.showGridLines=False
    today=datetime.date.today()

    D=_load_kb_data(); raw=D.get('calib_data',[])
    if not raw:
        banner(ws,"  📅 Vencimientos de Calibración",
               "  Sin datos — Carga el CSV con: python3 scripts/calib_from_csv.py --csv data_calib/calibraciones_meta.csv",9)
        msg=ws.cell(4,1,"⚠  No hay datos de calibración cargados. "
                    "Descarga el CSV de la card #75406 en Metabase y córrelo con calib_from_csv.py")
        msg.font=font(11,True,GOLD); msg.alignment=left
        ws.merge_cells(start_row=4,end_row=4,start_column=1,end_column=9)
        print("📅 Calibraciones: sin datos — cargar CSV"); return

    # ── Normalizar filas ──
    items=[]
    for r in raw:
        mes,anio=_mes_anio(_get(r,'mes','fecha','periodo'))
        if not mes: continue
        desc=_get(r,'descripcion','desc') or ''
        categoria=_get(r,'categoria','categoría','clase') or ''
        tipo=_tipo_equipo(desc, categoria)
        cant=int(_to_num(_get(r,'cantidad_equipos','cantidad','count')))
        vu=_to_num(_get(r,'valor_unitario','valor_unit','unitario'))
        codigo=_get(r,'codigo','código','sku') or ''
        items.append(dict(anio=anio,mes=mes,tipo=tipo,categoria=str(categoria),
                          codigo=str(codigo),desc=str(desc),vu=vu,cant=cant,
                          label=f"{_MESES_ES[mes-1]} {anio}"))

    items.sort(key=lambda x:(x['anio'],x['mes'],
               {'TC':0,'TP':1,'Medidor':2}.get(x['tipo'],9),x['codigo']))

    all_tipos=sorted({it['tipo'] for it in items},
                     key=lambda x:{'TC':0,'TP':1,'Medidor':2}.get(x,9))
    # períodos en orden
    seen=set(); periods=[]
    for it in items:
        k=(it['anio'],it['mes'])
        if k not in seen: seen.add(k); periods.append(k)
    P=len(periods); N=len(items)

    # IVA configurable (celda nombrada lógica): usamos 19% inline
    IVA_RATE=0.19

    TIPO_COL={'TC':TEAL,'TP':GOLD,'Medidor':'5B6BFF'}
    TIPO_LT ={'TC':TEAL_LT,'TP':GOLD_LT,'Medidor':'E8ECFF'}

    # ── Tipos de equipo distintos (por código) para la tablita resumen ──
    # Agrupa por código real (5496, 5497…) → cada SKU/equipo es una fila.
    equipos=[]; vistos=set()
    for it in items:
        k=it['codigo'] or it['desc']
        if k in vistos: continue
        vistos.add(k)
        equipos.append(dict(codigo=it['codigo'],desc=it['desc'],
                            tipo=it['tipo'],categoria=it['categoria'],vu=it['vu']))
    NE=len(equipos)

    # ── Layout determinístico (para referencias de fórmulas) ──
    # Pivot mensual: header=3, datos 4..3+P, total=4+P
    pv_hdr=3; pv_data0=4; pv_total=4+P
    # Tablita por tipo de equipo: título, header, datos, total
    bt_title=pv_total+2
    bt_hdr=bt_title+1
    bt_d0=bt_hdr+1
    bt_d1=bt_d0+NE-1
    bt_total=bt_d1+1
    # Detalle: título, header, datos
    det_title=bt_total+2
    det_hdr=det_title+1
    det_d0=det_hdr+1
    det_d1=det_d0+N-1
    # Columnas detalle: A=Mes B=Tipo C=Cat D=Cod E=Desc F=VU G=Cant H=Sub I=IVA J=Total
    A=f"$A${det_d0}:$A${det_d1}"   # label mes
    Bt=f"$B${det_d0}:$B${det_d1}"  # tipo
    Dc=f"$D${det_d0}:$D${det_d1}"  # código
    Gc=f"$G${det_d0}:$G${det_d1}"  # cantidad
    Hs=f"$H${det_d0}:$H${det_d1}"  # subtotal
    Ii=f"$I${det_d0}:$I${det_d1}"  # iva
    Jt=f"$J${det_d0}:$J${det_d1}"  # total c/IVA

    # ════ SECCIÓN 1 · RESUMEN MENSUAL (con fórmulas SUMIFS) ════
    pcols=[("Mes / Año",16)]
    for t in all_tipos: pcols+=[(f"{t}  N°",9),(f"{t}  Costo c/IVA",16)]
    pcols+=[("TOTAL Equipos",13),("TOTAL Costo c/IVA",18)]
    ncols=len(pcols)

    banner(ws,"  📅 Vencimientos de Calibración por Tipo de Equipo",
           f"  {N} líneas · {P} meses · fórmulas vivas (SUMIFS al detalle) · Fuente: Metabase card #75406",
           ncols)
    for i,(n,w) in enumerate(pcols,1):
        ws.cell(pv_hdr,i,n); ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,ncols,pv_hdr)
    ci=2
    for t in all_tipos:
        tc=TIPO_COL.get(t,"888888")
        ws.cell(pv_hdr,ci).fill=fill(tc); ws.cell(pv_hdr,ci+1).fill=fill(tc); ci+=2

    # Mapa tipo → letras de columna (N° y Costo) en el pivot
    tipo_cols={}
    ci=2
    for t in all_tipos:
        tipo_cols[t]=(get_column_letter(ci),get_column_letter(ci+1)); ci+=2
    cntL=[tipo_cols[t][0] for t in all_tipos]
    cosL=[tipo_cols[t][1] for t in all_tipos]
    totEqL=get_column_letter(ncols-1); totCoL=get_column_letter(ncols)

    rr=pv_data0
    for (anio,mes) in periods:
        diff=(anio*12+mes)-(today.year*12+today.month)
        if diff<0:   rf,rt="F0F0F0","999999"
        elif diff==0:rf,rt=RED_LT,RED
        elif diff==1:rf,rt=GOLD_LT,GOLD
        else:        rf,rt=None,TXT
        lbl=f"{_MESES_ES[mes-1]} {anio}"
        mc=ws.cell(rr,1,lbl); mc.font=font(10,True,rt); mc.alignment=center; mc.border=border_all
        if rf: mc.fill=fill(rf)
        ci=2
        for t in all_tipos:
            # N° equipos = SUMIFS(cantidad, mes=label, tipo=t)
            f_cnt=f'=SUMIFS({Gc},{A},$A{rr},{Bt},"{t}")'
            # Costo c/IVA = SUMIFS(total, mes=label, tipo=t)
            f_cos=f'=SUMIFS({Jt},{A},$A{rr},{Bt},"{t}")'
            for sc,(f_,nf) in enumerate([(f_cnt,'#,##0'),(f_cos,'#,##0')]):
                cc=ws.cell(rr,ci+sc,f_); cc.border=border_all
                cc.font=font(10,False,rt); cc.alignment=center; cc.number_format=nf
                if rf: cc.fill=fill(rf)
                else:  cc.fill=fill(TIPO_LT.get(t,"F4F4F4"))
            ci+=2
        # TOTAL Equipos / Costo = suma de columnas de tipo en la fila
        f_teq="="+"+".join(f"{L}{rr}" for L in cntL)
        f_tco="="+"+".join(f"{L}{rr}" for L in cosL)
        for sc,(f_,nf) in enumerate([(f_teq,'#,##0'),(f_tco,'#,##0')]):
            cc=ws.cell(rr,ncols-1+sc,f_); cc.border=border_all
            cc.font=font(10,True,rt); cc.alignment=center; cc.number_format=nf
            if rf: cc.fill=fill(rf)
        rr+=1

    # Fila TOTAL = SUM de cada columna sobre los meses
    tc0=ws.cell(pv_total,1,"TOTAL"); tc0.fill=fill(NAVY); tc0.font=font(10,True,WHITE)
    tc0.alignment=center; tc0.border=border_all
    for c in range(2,ncols+1):
        L=get_column_letter(c)
        cc=ws.cell(pv_total,c,f"=SUM({L}{pv_data0}:{L}{pv_total-1})")
        cc.fill=fill(TEAL); cc.font=font(11 if c>=ncols-1 else 10,True,WHITE)
        cc.alignment=center; cc.border=border_all; cc.number_format='#,##0'

    # ════ SECCIÓN 1b · RESUMEN POR TIPO DE EQUIPO (tablita, fórmulas SUMIFS) ════
    # Una fila por código/SKU (TC Baja, TC Media, TP Media, Medidor…), agregando
    # TODOS los meses. Valor unitario = costo de calibrar 1 equipo de ese tipo.
    bcols=[("Código",9),("Tipo",10),("Equipo / Descripción",28),
           ("Valor Unit. COP",15),("Total Equipos",13),
           ("Subtotal COP",15),("IVA 19% COP",13),("Total c/IVA COP",16)]
    nbc=len(bcols)
    bh=ws.cell(bt_title,1,"  🔧 Resumen por Tipo de Equipo  ·  costo unitario de calibración × equipos a vencer (todos los meses)")
    bh.fill=fill(NAVY); bh.font=font(11,True,WHITE); bh.alignment=left
    ws.merge_cells(start_row=bt_title,end_row=bt_title,start_column=1,end_column=nbc)
    for c in range(1,nbc+1): ws.cell(bt_title,c).border=border_all
    for i,(n,w) in enumerate(bcols,1):
        ws.cell(bt_hdr,i,n)
        ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,nbc,bt_hdr)

    rr=bt_d0
    for e in equipos:
        cod=e['codigo']
        ws.cell(rr,1,cod).alignment=center
        ct=ws.cell(rr,2,e['tipo'])
        ws.cell(rr,3,e['desc']).alignment=left
        cvu=ws.cell(rr,4,e['vu']); cvu.number_format='#,##0'; cvu.alignment=center
        # Total equipos a vencer = SUMIFS(cantidad detalle, código = este código)
        cte=ws.cell(rr,5,f'=SUMIFS({Gc},{Dc},$A{rr})'); cte.number_format='#,##0'; cte.alignment=center
        # Subtotal/IVA/Total = SUMIFS al detalle por código
        cs=ws.cell(rr,6,f'=SUMIFS({Hs},{Dc},$A{rr})'); cs.number_format='#,##0'; cs.alignment=center
        cv=ws.cell(rr,7,f'=SUMIFS({Ii},{Dc},$A{rr})'); cv.number_format='#,##0'; cv.alignment=center
        cto=ws.cell(rr,8,f'=SUMIFS({Jt},{Dc},$A{rr})'); cto.number_format='#,##0'; cto.alignment=center
        for c in range(1,nbc+1):
            cc=ws.cell(rr,c); cc.border=border_all
            if c==2:
                cc.fill=fill(TIPO_LT.get(e['tipo'],"F4F4F4")); cc.font=font(10,True,TIPO_COL.get(e['tipo'],TXT)); cc.alignment=center
            else:
                cc.font=font(10)
        rr+=1
    # Fila TOTAL de la tablita
    tb0=ws.cell(bt_total,1,"TOTAL"); tb0.fill=fill(NAVY); tb0.font=font(10,True,WHITE)
    tb0.alignment=center; tb0.border=border_all
    ws.merge_cells(start_row=bt_total,end_row=bt_total,start_column=1,end_column=3)
    for c in range(2,4): ws.cell(bt_total,c).border=border_all
    for c in (5,6,7,8):  # Total equipos, subtotal, iva, total c/IVA
        L=get_column_letter(c)
        cc=ws.cell(bt_total,c,f"=SUM({L}{bt_d0}:{L}{bt_d1})")
        cc.fill=fill(TEAL); cc.font=font(11 if c==8 else 10,True,WHITE)
        cc.alignment=center; cc.border=border_all; cc.number_format='#,##0'
    ws.cell(bt_total,4).fill=fill(NAVY); ws.cell(bt_total,4).border=border_all  # valor unit no se suma

    # ════ SECCIÓN 2 · DETALLE POR PRODUCTO (con fórmulas) ════
    dcols=[("Mes / Año",16),("Tipo",10),("Categoría",11),("Código",9),
           ("Descripción",28),("Valor Unit. COP",15),("Cantidad",10),
           ("Subtotal COP",15),("IVA 19% COP",13),("Total c/IVA COP",16)]
    ht=ws.cell(det_title,1,"  📋 Detalle por Producto y Mes  ·  Subtotal=Valor×Cant · IVA=Subtotal×19% · Total=Subtotal+IVA")
    ht.fill=fill(NAVY); ht.font=font(11,True,WHITE); ht.alignment=left
    ws.merge_cells(start_row=det_title,end_row=det_title,start_column=1,end_column=len(dcols))
    for c in range(1,len(dcols)+1): ws.cell(det_title,c).border=border_all
    for i,(n,w) in enumerate(dcols,1):
        ws.cell(det_hdr,i,n)
        if i>ncols: ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,len(dcols),det_hdr)

    rr=det_d0
    for it in items:
        ws.cell(rr,1,it['label']).alignment=center
        ws.cell(rr,2,it['tipo'])
        ws.cell(rr,3,it['categoria']).alignment=left
        ws.cell(rr,4,it['codigo']).alignment=center
        ws.cell(rr,5,it['desc']).alignment=left
        cvu=ws.cell(rr,6,it['vu']); cvu.number_format='#,##0'; cvu.alignment=center
        cct=ws.cell(rr,7,it['cant']); cct.number_format='#,##0'; cct.alignment=center
        # Fórmulas:
        csub=ws.cell(rr,8,f"=F{rr}*G{rr}"); csub.number_format='#,##0'; csub.alignment=center
        civa=ws.cell(rr,9,f"=ROUND(H{rr}*{IVA_RATE},0)"); civa.number_format='#,##0'; civa.alignment=center
        ctot=ws.cell(rr,10,f"=H{rr}+I{rr}"); ctot.number_format='#,##0'; ctot.alignment=center
        for c in range(1,11):
            cc=ws.cell(rr,c); cc.border=border_all
            if c==2:
                cc.fill=fill(TIPO_LT.get(it['tipo'],"F4F4F4")); cc.font=font(10,True,TIPO_COL.get(it['tipo'],TXT)); cc.alignment=center
            else:
                cc.font=font(10)
        rr+=1
    if det_d1>=det_d0:
        tab=Table(displayName="DetCalib",ref=f"A{det_hdr}:{get_column_letter(len(dcols))}{det_d1}")
        tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
        ws.add_table(tab)

    ws.freeze_panes="A4"
    nota=("🔴 Mes actual · 🟡 Próximo mes · ⬜ Futuro · ▦ Pasado (gris)  ·  "
          "Resumen arriba = SUMIFS al detalle · Detalle = fórmulas Subtotal/IVA/Total  ·  "
          "Fuente: Metabase card #75406 (IVA 19%)")
    ws.cell(det_d1+2,1,nota).font=font(9,False,"7A8094")
    ws.merge_cells(start_row=det_d1+2,end_row=det_d1+2,start_column=1,end_column=max(ncols,len(dcols)))
    print(f"📅 Calibraciones OK (con fórmulas): {P} meses · {N} líneas · detalle filas {det_d0}-{det_d1}")



# ════════════════════════════════════════════════════════════════
# HOJA 🚚 — LEAD TIMES
# Fuente de los días: provider_responses.delivery_days (real).
# _LT_META solo describe categoría/país/condiciones — NUNCA días.
# ════════════════════════════════════════════════════════════════

# {clave_norm: (tipo, pais, entrega, stock_inm, ped_min, notas)}
# Clave = fragmento del nombre (minúsculas, sin sufijos legales)
_LT_META = {
    "selda":                    ("Medidores/TC/TP",     "Colombia","Nacional",  "Parcial",  "1 unidad",     "Certificado CIDET. Stock parcial TC y medidores."),
    "proelco":                  ("Distribuidor",        "Colombia","Nacional",  "Sí",       "Sin mínimo",   "Bodega local. Entrega inmediata frecuente."),
    "adler":                    ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Distribuidor con catálogo amplio."),
    "disico":                   ("Fabricante TC/TP",    "Colombia","Nacional",  "Parcial",  "5 unidades",   "Fabricante TCs. Tiempo varía por especificación."),
    "laumayer":                 ("Fabricante Tableros", "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Condensadores y tableros eléctricos."),
    "acj":                      ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Distribución materiales eléctricos."),
    "inpel":                    ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "1 unidad",     "Distribución eléctrica nacional."),
    "disproel":                 ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Distribución materiales eléctricos."),
    "vaelectricos":             ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Materiales eléctricos."),
    "jimaco":                   ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Distribución eléctrica."),
    "fisa":                     ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Materiales y equipos eléctricos."),
    "incomelec":                ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Distribución eléctrica."),
    "ectricol":                 ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Distribución materiales eléctricos."),
    "grupo defa":               ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Materiales eléctricos."),
    "red + electric":           ("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Materiales eléctricos."),
    "suministros automatizados":("Distribuidor",        "Colombia","Nacional",  "Parcial",  "Sin mínimo",   "Suministros industriales."),
    "metrobit":                 ("Importador AMI",      "Europa",  "Importado", "Parcial",  "1 unidad",     "Medidores inteligentes. Representante local."),
    "emsi":                     ("Importador AMI",      "China",   "Importado", "No",       "5 unidades",   "Marca MICROSTAR."),
    "hexing":                   ("Importador AMI",      "China",   "Importado", "No",       "50 unidades",  "Medidores AMI. LC + flete marítimo."),
    "mn technologies":          ("Importador AMI",      "China",   "Importado", "No",       "5 unidades",   "LC 30 días."),
    "electrocables":            ("Cables/Conductores",  "Colombia","Nacional",  "Parcial",  "50 m/bobina",  "Conductores eléctricos."),
    "prysmian":                 ("Cables/Conductores",  "Italia",  "Importado", "Parcial",  "50 m/bobina",  "Distribución local Colombia."),
    "ja solar":                 ("Solar/Renovables",    "China",   "Importado", "No",       "Pallet",       "Paneles fotovoltaicos. LC 45 días + flete."),
    "growatt":                  ("Solar/Renovables",    "China",   "Importado", "No",       "1 unidad",     "Inversores. Representante local."),
    "il sole":                  ("Ingeniería Solar",    "Colombia","Nacional",  "N/A",      "Por proyecto", "Proyectos solares llave en mano."),
    "solcity":                  ("Ingeniería Solar",    "Colombia","Nacional",  "N/A",      "Por proyecto", "Proyectos solares llave en mano."),
    "4s":                       ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería eléctrica integral."),
    "dce":                      ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Instalaciones eléctricas industriales."),
    "noatec":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Sistemas de medición y control."),
    "tevium":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Proyectos AMI/Smart Grid."),
    "keb":                      ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería eléctrica."),
    "faccel":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Proyectos eléctricos."),
    "ryctel":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Automatización y control."),
    "eqysol":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Proyectos eléctricos industriales."),
    "smart projects":           ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería y proyectos especiales."),
    "m y d":                    ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería y diseño eléctrico."),
    "jlring":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería eléctrica."),
    "idosde":                   ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería eléctrica."),
    "k&v":                      ("Ingeniería",          "Colombia","Nacional",  "N/A",      "Por proyecto", "Ingeniería y diseño."),
    "instalaciones y soluciones":("Ingeniería",         "Colombia","Nacional",  "N/A",      "Por proyecto", "Instalaciones eléctricas."),
    "tableros electricos":      ("Fabricante Tableros", "Colombia","Nacional",  "No",       "Por proyecto", "Tableros a medida."),
    "industrias rebra":         ("Fabricante Tableros", "Colombia","Nacional",  "No",       "Por proyecto", "Tableros y paneles a medida."),
    "servimeters":              ("Calibración",         "Colombia","Nacional",  "N/A",      "Por servicio", "Calibración ISO 17025. Tiempo según laboratorio."),
    "veritest":                 ("Calibración",         "Colombia","Nacional",  "N/A",      "Por servicio", "Ensayos y certificaciones."),
    "certeam":                  ("Calibración",         "Colombia","Nacional",  "N/A",      "Por servicio", "Certificaciones técnicas."),
    "rentek":                   ("Arriendo Operativo",  "Colombia","Nacional",  "Parcial",  "1 activo",     "Disponibilidad según flota."),
    "renting colombia":         ("Arriendo Operativo",  "Colombia","Nacional",  "Parcial",  "1 unidad",     "Arriendo de equipos eléctricos."),
    "formas logisticas":        ("Logística",           "Colombia","Nacional",  "Sí",       "Sin mínimo",   "Operador logístico. Entrega 24-48h."),
    "operador de transporte":   ("Logística",           "Colombia","Nacional",  "Sí",       "Por viaje",    "Transporte terrestre y marítimo."),
    "erasmus":                  ("Consultoría",         "Colombia","Nacional",  "N/A",      "Por servicio", "Normativa y consultoría."),
    "sdt":                      ("Consultoría",         "Colombia","Nacional",  "N/A",      "Por servicio", "Consultoría técnica especializada."),
    "gers":                     ("Consultoría",         "Colombia","Nacional",  "N/A",      "Por servicio", "Consultoría sistemas de energía."),
    "normarh":                  ("Consultoría",         "Colombia","Nacional",  "N/A",      "Por servicio", "Normativa eléctrica."),
    "corp centro":              ("Consultoría",         "Colombia","Nacional",  "N/A",      "Por servicio", "Centro de innovación sector eléctrico."),
}

def _lt_meta(name):
    """Devuelve (tipo, pais, entrega, stock, ped_min, notas) desde _LT_META."""
    n=(name or "").lower()
    for s in (" sas"," s.a.s."," s.a.s"," s.a."," ltda"," technologies colombia"," colombia"," ingenieria"," 1 "):
        n=n.replace(s," ")
    n=n.strip()
    if n in _LT_META: return _LT_META[n]
    for k,v in _LT_META.items():
        if k and k in n: return v
    for k,v in _LT_META.items():
        if k and len(n)>3 and n in k: return v
    return ("Por clasificar","Colombia","Nacional","Por confirmar","Por confirmar","")

def _lt_days(name, kb_lt):
    """Busca el lead time real en kb_lt por nombre normalizado."""
    if not kb_lt: return None
    n=_norm(name)
    for k,v in kb_lt.items():
        if _norm(k)==n: return v
    for k,v in kb_lt.items():
        nk=_norm(k)
        if nk and (nk in n or n in nk): return v
    return None

def build_lead_times(wb):
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    title="🚚 Lead Times"
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    ws.sheet_view.showGridLines=False

    D=_load_kb_data()
    kb_lt=D.get('lead_times',{})        # {prov_name: avg_days}  ← de provider_responses.delivery_days
    kb_cnt=D.get('lead_times_count',{}) # {prov_name: n_responses}
    offers=D.get('offers_by_sku',{})
    provs=_load_providers()

    # SKU count per normalized provider
    prov_sku=defaultdict(set)
    for sid,offs in offers.items():
        for o in offs:
            if o.get('price'): prov_sku[_norm(o['provider'])].add(sid)
    def sku_cnt(name):
        n=_norm(name); best=0
        for pn,sids in prov_sku.items():
            if pn==n or (pn and n and (pn in n or n in pn)):
                best=max(best,len(sids))
        return best

    def resp_cnt(name):
        n=_norm(name)
        for k,v in kb_cnt.items():
            if _norm(k)==n or (n and _norm(k) and (_norm(k) in n or n in _norm(k))):
                return v
        return 0

    # Build row list: providers.json como fuente canónica
    seen=set(); all_rows=[]
    for p in provs:
        emp=p.get('empresa','')
        if not emp or emp.upper()=='PRUEBA': continue
        n=_norm(emp)
        if n in seen: continue
        seen.add(n)
        days=_lt_days(emp,kb_lt)
        meta=_lt_meta(emp)
        all_rows.append((emp.upper(),meta[0],meta[1],meta[2],
                         days,resp_cnt(emp),meta[3],meta[4],meta[5]))

    # Añadir cualquier proveedor en kb_lt no cubierto por providers.json
    for pname in kb_lt:
        if any(x in pname.upper() for x in ('PRUEB','TEST')): continue
        n=_norm(pname)
        if any(n==sn or (n and sn and (n in sn or sn in n)) for sn in seen): continue
        seen.add(n)
        days=kb_lt[pname]
        meta=_lt_meta(pname)
        all_rows.append((pname.upper(),meta[0],meta[1],meta[2],
                         days,resp_cnt(pname),meta[3],meta[4],meta[5]))

    # Ordenar: con dato real primero, luego por tipo, luego alfabético
    all_rows.sort(key=lambda r:(0 if r[4] is not None else 1, r[1], r[0]))
    N=len(all_rows)

    # Distinct tipos (first-appearance)
    seen_t=set(); all_tipos=[]
    for r in all_rows:
        t=r[1]
        if t not in seen_t: seen_t.add(t); all_tipos.append(t)
    T=len(all_tipos)

    NCOLS=12

    # ── Layout determinístico ──
    sum_title_row=3
    pv_hdr=4; pv_data0=5; pv_total=5+T
    det_title=pv_total+2
    det_hdr=det_title+1
    det_d0=det_hdr+1
    det_d1=det_d0+N-1

    Btyp=f"$B${det_d0}:$B${det_d1}"  # tipo
    Eprom=f"$E${det_d0}:$E${det_d1}" # LT real (numérico o vacío)

    n_real=sum(1 for r in all_rows if r[4] is not None)
    banner(ws,"  🚚 Lead Times por Proveedor",
           f"  {N} proveedores · {n_real} con dato real (delivery_days) · {N-n_real} sin dato aún · jun-2026",
           NCOLS)

    # ════ SECCIÓN 1 — RESUMEN POR TIPO ════
    st=ws.cell(sum_title_row,1,"  📊 Resumen por Tipo  ·  AVERAGEIF / COUNTIFS sobre col E (LT Real)  ·  ⏳ = sin pedidos aún")
    st.fill=fill(NAVY); st.font=font(11,True,WHITE); st.alignment=left
    ws.merge_cells(start_row=sum_title_row,end_row=sum_title_row,start_column=1,end_column=NCOLS)
    for c in range(1,NCOLS+1): ws.cell(sum_title_row,c).border=border_all

    SUM_COLS=[("Tipo de Proveedor",22),("# Total",10),
              ("# Con LT Real",14),("LT Promedio Real (días)",20)]
    for i,(n,w) in enumerate(SUM_COLS,1):
        ws.cell(pv_hdr,i,n); ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,len(SUM_COLS),pv_hdr)

    rr=pv_data0
    for t in all_tipos:
        ws.cell(rr,1,t).alignment=left; ws.cell(rr,1).font=font(10); ws.cell(rr,1).border=border_all
        f_tot  =f'=COUNTIF({Btyp},A{rr})'
        f_real =f'=COUNTIFS({Btyp},A{rr},{Eprom},">0")'
        f_avg  =f'=IFERROR(ROUND(AVERAGEIF({Btyp},A{rr},{Eprom}),1),"Sin dato")'
        for ci,(f_,nf) in enumerate([(f_tot,'#,##0'),(f_real,'#,##0'),(f_avg,'0.0')],2):
            cc=ws.cell(rr,ci,f_); cc.border=border_all
            cc.font=font(10); cc.alignment=center; cc.number_format=nf
        rr+=1

    # Fila TOTAL
    tc=ws.cell(pv_total,1,"TOTAL")
    tc.fill=fill(NAVY); tc.font=font(10,True,WHITE); tc.alignment=center; tc.border=border_all
    for ci,(f_,nf) in enumerate([
        (f'=SUM(B{pv_data0}:B{pv_total-1})','#,##0'),
        (f'=SUM(C{pv_data0}:C{pv_total-1})','#,##0'),
        (f'=IFERROR(ROUND(AVERAGEIF({Eprom},">0"),1),"Sin dato")','0.0'),
    ],2):
        cc=ws.cell(pv_total,ci,f_); cc.fill=fill(TEAL)
        cc.font=font(10,True,WHITE); cc.alignment=center; cc.border=border_all; cc.number_format=nf

    # ════ SECCIÓN 2 — DETALLE POR PROVEEDOR ════
    # Col E = LT Real (numérico o vacío); Col H = Velocidad (fórmula); Col K = Estado
    DCOLS=[("Proveedor",28),("Tipo",20),("País",12),("Entrega",12),
           ("LT Real (días)",15),("# Respuestas",13),
           ("Velocidad",22),("Stock Inm.",12),("Ped. Mínimo",14),
           ("# SKUs BIA",11),("Estado Dato",14),("Notas",36)]

    dt=ws.cell(det_title,1,"  📋 Detalle por Proveedor  ·  Col E = promedio real de delivery_days · Verde=Nacional · Rojo=Importado")
    dt.fill=fill(NAVY); dt.font=font(11,True,WHITE); dt.alignment=left
    ws.merge_cells(start_row=det_title,end_row=det_title,start_column=1,end_column=NCOLS)
    for c in range(1,NCOLS+1): ws.cell(det_title,c).border=border_all

    for i,(n,w) in enumerate(DCOLS,1):
        ws.cell(det_hdr,i,n); ws.column_dimensions[get_column_letter(i)].width=w
    style_header(ws,len(DCOLS),det_hdr)

    rr=det_d0
    for prov,tipo,pais,entrega,days,nresp,stock,ped_min,notas in all_rows:
        skus=sku_cnt(prov)
        real=days is not None

        ws.cell(rr,1,prov).alignment=left
        ws.cell(rr,2,tipo).alignment=left
        ws.cell(rr,3,pais).alignment=center
        cen=ws.cell(rr,4,entrega); cen.alignment=center
        # Col E: numérico si tenemos dato real, vacío si no
        if real:
            cp=ws.cell(rr,5,days); cp.alignment=center; cp.number_format='#,##0'
        else:
            ws.cell(rr,5,None)
        # Col F: # respuestas
        ws.cell(rr,6,nresp if nresp else None).alignment=center
        # Col G: fórmula Velocidad — evalúa E directamente
        fv=f'=IF(E{rr}="","⏳ Sin dato",IF(E{rr}<=5,"🟢 Muy rápido",IF(E{rr}<=15,"🟡 Medio","🔴 Lento")))'
        ws.cell(rr,7,fv).alignment=center
        ws.cell(rr,8,stock).alignment=center
        ws.cell(rr,9,ped_min).alignment=left
        ws.cell(rr,10,skus if skus>0 else "—").alignment=center
        # Col K: estado visible
        est="✅ Real" if real else "⏳ Por consultar"
        ck=ws.cell(rr,11,est); ck.alignment=center
        ws.cell(rr,12,notas).alignment=left

        for ci in range(1,13):
            cc=ws.cell(rr,ci); cc.border=border_all
            if not cc.font.bold: cc.font=font(10)
        # Color Entrega
        if entrega=="Nacional":
            cen.fill=fill(GREEN_LT); cen.font=font(10,True,GREEN)
        elif entrega=="Importado":
            cen.fill=fill(RED_LT); cen.font=font(10,False,RED)
        # Color Estado
        if real:
            ck.fill=fill(GREEN_LT); ck.font=font(10,True,GREEN)
        else:
            ck.fill=fill(GOLD_LT); ck.font=font(10,False,GOLD)
        rr+=1

    if det_d1>=det_d0:
        tab=Table(displayName="LeadTimes",ref=f"A{det_hdr}:{get_column_letter(NCOLS)}{det_d1}")
        tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
        ws.add_table(tab)

    # Color scale solo sobre celdas con dato (blancos quedan sin color)
    ws.conditional_formatting.add(f"E{det_d0}:E{det_d1}",
        ColorScaleRule(start_type='min',start_color=GREEN_LT,
                       mid_type='percentile',mid_value=50,mid_color=GOLD_LT,
                       end_type='max',end_color=RED_LT))
    ws.conditional_formatting.add(f"H{det_d0}:H{det_d1}",
        CellIsRule(operator='equal',formula=['"Sí"'],fill=fill(GREEN_LT),font=font(10,True,GREEN)))
    ws.conditional_formatting.add(f"H{det_d0}:H{det_d1}",
        CellIsRule(operator='equal',formula=['"No"'],fill=fill(RED_LT),font=font(10,False,RED)))
    ws.conditional_formatting.add(f"H{det_d0}:H{det_d1}",
        CellIsRule(operator='equal',formula=['"Parcial"'],fill=fill(GOLD_LT),font=font(10,False,GOLD)))

    ws.freeze_panes="A5"

    nota=(f"✅ Verde = LT calculado del promedio real de tus pedidos (delivery_days en provider_responses)  ·  "
          f"⏳ Amarillo = sin respuestas reales aún  ·  "
          f"Velocidad: 🟢≤5 · 🟡6-15 · 🔴>15 días  ·  jun-2026")
    ws.cell(det_d1+2,1,nota).font=font(9,False,"7A8094")
    ws.merge_cells(start_row=det_d1+2,end_row=det_d1+2,start_column=1,end_column=NCOLS)
    print(f"🚚 Lead Times OK: {N} proveedores · {n_real} con LT real · {N-n_real} sin dato · filas {det_d0}-{det_d1}")


# ─── REGISTRO DE HOJAS ────────────────────────────────────────────
BUILDERS={
    "operador":    build_operador,
    "condiciones": build_condiciones,
    "historico":   build_historico,
    "calibracion": build_calibracion,
    "lead_times":  build_lead_times,
}

def reorder_and_color(wb):
    wb._sheets.sort(key=lambda s: TAB_ORDER.index(s.title) if s.title in TAB_ORDER else 99)
    for s in wb._sheets:
        if s.title in TAB_COLORS: s.sheet_properties.tabColor=TAB_COLORS[s.title]

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--sheet",required=True,choices=list(BUILDERS.keys()),help="Hoja a agregar/reemplazar")
    ap.add_argument("--file",default=str(ROOT/"KB_BIA_Proveedores.xlsx"))
    args=ap.parse_args()

    path=Path(args.file)
    if not path.exists():
        print(f"❌ No existe {path}. Genera primero el Excel completo con build_excel.py")
        return
    wb=load_workbook(path)
    BUILDERS[args.sheet](wb)
    reorder_and_color(wb)
    wb.save(path)
    print(f"✓ Guardado: {path}")
    print("Hojas:",[s.title for s in wb._sheets])

if __name__=="__main__":
    main()
