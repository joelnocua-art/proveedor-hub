#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_calibracion_doble.py — Genera un .xlsx con DOS hojas:
  1. 📅 Calibraciones   — resumen mensual + tabla por tipo + detalle (datos agregados)
  2. 📋 Listado Equipos — un equipo por fila con serial, SKU, fecha exacta, precio unitario

Uso:
  python3 scripts/build_calibracion_doble.py \
      --csv-agg  data_calib/calibraciones_meta.csv \
      --csv-det  <ruta al CSV granular (726 filas)> \
      --out      Calibraciones.xlsx
"""
import argparse, csv, json, sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(Path(__file__).parent))

import add_sheet as A
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill


# ── helpers de paleta ──────────────────────────────────────────
ESTADO_COLOR = {
    '2026-06': (A.RED_LT,  A.RED),
    '2026-07': (A.GOLD_LT, A.GOLD),
}
TIPO_COL = {'TC': A.TEAL, 'TP': A.GOLD, 'Medidor': '5B6BFF'}
TIPO_LT  = {'TC': A.TEAL_LT, 'TP': A.GOLD_LT, 'Medidor': 'E8ECFF'}

_MESES_ES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

def mes_label(yyyymm):
    """'2026-07' → 'Julio 2026'"""
    try:
        y, m = yyyymm.split('-')
        return f"{_MESES_ES[int(m)-1]} {y}"
    except Exception:
        return yyyymm

def tipo_from_desc(desc, cat=''):
    d = (desc or '').upper()
    if d.startswith('TC') or ' TC' in d: return 'TC'
    if d.startswith('TP') or ' TP' in d: return 'TP'
    if 'MEDIDOR' in d or 'MEDID' in d:   return 'Medidor'
    if 'MEDIDOR' in (cat or '').upper():  return 'Medidor'
    return d.split()[0][:10] if d else 'Otro'

def tonum(s):
    return float(str(s).replace(',','').replace('$','').replace(' ','').strip() or 0)


# ══════════════════════════════════════════════════════════════════
# HOJA 1 — resumen mensual (mismo build_calibracion existente)
# ══════════════════════════════════════════════════════════════════
def _agregar_desde_detalle(det_csv):
    """Agrega el CSV granular (1 fila/serial) al formato que espera
    build_calibracion: 1 fila por (mes, código). Garantiza que el total
    de la Hoja 1 cuadre exactamente con el de la Hoja 2."""
    import collections
    grupos = collections.OrderedDict()  # (mes, codigo) -> dict acumulado
    with open(det_csv, newline='', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            mes = (r.get('mes_vencimiento') or '').strip()
            cod = (r.get('codigo') or '').strip()
            if not mes or not cod:
                continue
            k = (mes, cod)
            g = grupos.get(k)
            if g is None:
                grupos[k] = {
                    "mes": mes,
                    "categoria": r.get('categoria', ''),    # Media / Baja / Medidor
                    "codigo": cod,
                    "descripcion": r.get('descripcion', ''),  # "TC Media Tensión", etc.
                    "valor_unitario": tonum(r.get('valor_unitario', 0)),
                    "cantidad_equipos": 1,
                }
            else:
                g["cantidad_equipos"] += 1
    return list(grupos.values())


def build_hoja1(wb, det_csv):
    # Se agrega DESDE el detalle granular → misma fuente que la Hoja 2
    calib = _agregar_desde_detalle(det_csv)
    A._load_kb_data = lambda: {"calib_data": calib}
    A.build_calibracion(wb)
    ws = wb["📅 Calibraciones"]
    ws.sheet_properties.tabColor = A.TAB_COLORS.get("📅 Calibraciones", "009688")
    total = sum(int(c["cantidad_equipos"]) for c in calib)
    print(f"📅 Hoja 1 agregada desde el detalle: {total} equipos en {len(calib)} líneas")


# ══════════════════════════════════════════════════════════════════
# HOJA 2 — Listado granular por equipo/serial
# ══════════════════════════════════════════════════════════════════
def build_hoja2(wb, det_csv):
    title = "📋 Listado Equipos"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    # Leer CSV
    rows = []
    with open(det_csv, newline='', encoding='utf-8-sig') as f:
        for r in csv.DictReader(f):
            rows.append(r)

    N = len(rows)

    # Nota al pie: ambas hojas salen del mismo detalle → totales cuadran
    NOTA_DISC = (
        f"Listado fila-por-serial · {N} equipos · El resumen de la Hoja 1 se agrega desde este mismo "
        "detalle, por lo que los totales de ambas hojas coinciden exactamente."
    )

    COLS = [
        ("N°",        5),
        ("Serial",   18),
        ("Código",    8),
        ("Tipo",     10),
        ("Categoría",11),
        ("SKU / Equipo",       42),
        ("Descripción Servicio",20),
        ("Vencimiento",      14),
        ("Mes Vence",        13),
        ("Valor Unit. COP",  15),
        ("IVA 19% COP",      13),
        ("Total c/IVA COP",  16),
        ("Certificado URL",  40),
    ]
    NC = len(COLS)

    # ── Banner ──
    A.banner(ws,
        "  📋 Listado de Equipos a Vencer — Detalle por Serial",
        f"  {N} equipos · Jun-Dic 2026 · fila por serial · precio unitario de calibración",
        NC)

    # ── KPIs fila 3 ──
    hdr = 4; d0 = 5; d1 = d0 + N - 1
    # KPIs como fórmulas vivas sobre el detalle (B=serial, C=código)
    kpis = [
        ("Total equipos",   f"=COUNTA(B{d0}:B{d1})"),
        ("TC Baja (5496)",  f'=COUNTIF(C{d0}:C{d1},"5496")'),
        ("TC Media (5498)", f'=COUNTIF(C{d0}:C{d1},"5498")'),
        ("TP Media (5497)", f'=COUNTIF(C{d0}:C{d1},"5497")'),
        ("Medidor (6024)",  f'=COUNTIF(C{d0}:C{d1},"6024")'),
    ]
    col = 1
    for lbl, val in kpis:
        lc = ws.cell(3, col, lbl)
        lc.fill = A.fill(A.NAVY); lc.font = A.font(9, True, A.WHITE)
        lc.alignment = A.center; lc.border = A.border_all
        vc = ws.cell(3, col + 1, val)
        vc.fill = A.fill(A.TEAL); vc.font = A.font(12, True, A.WHITE)
        vc.alignment = A.center; vc.border = A.border_all
        vc.number_format = '#,##0'
        col += 2
    for c in range(col, NC + 1):
        ws.cell(3, c).fill = A.fill(A.GREY_LT); ws.cell(3, c).border = A.border_all
    ws.row_dimensions[3].height = 22

    # ── Encabezado ──
    for i, (name, w) in enumerate(COLS, 1):
        ws.cell(hdr, i, name)
        ws.column_dimensions[get_column_letter(i)].width = w
    A.style_header(ws, NC, hdr)

    # ── Datos ──
    import datetime
    today = datetime.date.today()
    for idx, r in enumerate(rows, 1):
        rr = d0 + idx - 1
        mes = r.get('mes_vencimiento', '')
        dia = r.get('dia_vencimiento', '')
        tipo = tipo_from_desc(r.get('descripcion',''), r.get('categoria',''))
        vu = tonum(r.get('valor_unitario', 0))
        # IVA y Total como FÓRMULAS vivas (J=Valor Unit, K=IVA, L=Total)
        iva = f"=ROUND(J{rr}*0.19,0)"
        tot = f"=J{rr}+K{rr}"

        # Urgencia: solo TEXTO de color en la fecha (no se rellena la fila)
        try:
            y, m = mes.split('-')
            diff = (int(y)*12 + int(m)) - (today.year*12 + today.month)
        except Exception:
            diff = 99
        if diff <= 0:   fecha_color, fecha_bold = A.RED,  True   # vence este mes o antes
        elif diff == 1: fecha_color, fecha_bold = A.GOLD, True   # próximo mes
        else:           fecha_color, fecha_bold = A.TXT,  False

        vals = [
            (idx,         'center', None),
            (r.get('serial',''), 'left', None),
            (r.get('codigo',''), 'center', None),
            (tipo,        'center', None),
            (r.get('categoria',''), 'center', None),
            (r.get('sku',''),         'left', None),   # SKU / Equipo (detalle real)
            (r.get('descripcion',''), 'left', None),   # Descripción del servicio
            (dia,         'center', None),
            (mes_label(mes), 'center', None),
            (vu,          'center', '#,##0'),
            (iva,         'center', '#,##0'),
            (tot,         'center', '#,##0'),
            (r.get('certificado_url',''), 'left', None),
        ]
        for ci, (val, align, extra) in enumerate(vals, 1):
            cc = ws.cell(rr, ci, val)
            cc.alignment = A.center if align == 'center' else A.left
            cc.border = A.border_all
            cc.font = A.font(9)
            if ci == 4:  # Tipo — chip de color suave
                cc.fill = A.fill(TIPO_LT.get(tipo, A.GREY_LT))
                cc.font = A.font(9, True, TIPO_COL.get(tipo, A.TXT))
            elif ci in (8, 9):  # Vencimiento / Mes Vence — texto de urgencia
                cc.font = A.font(9, fecha_bold, fecha_color)
            if isinstance(extra, str) and extra:
                cc.number_format = extra

        # Hipervínculo certificado (col 13)
        url = r.get('certificado_url', '')
        if url and url.startswith('http'):
            link_cell = ws.cell(rr, 13)
            link_cell.hyperlink = url
            link_cell.font = A.font(9, False, '1A6FB8')

    # ── Tabla ──
    if d1 >= d0:
        tab = Table(displayName="ListadoEquipos",
                    ref=f"A{hdr}:{get_column_letter(NC)}{d1}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showRowStripes=True)
        ws.add_table(tab)

    ws.freeze_panes = f"A{d0}"

    # ── Fila TOTAL (fórmulas SUM) debajo de la tabla ──
    tr = d1 + 1
    tc0 = ws.cell(tr, 1, "TOTAL")
    tc0.fill = A.fill(A.NAVY); tc0.font = A.font(10, True, A.WHITE)
    tc0.alignment = A.center; tc0.border = A.border_all
    ws.merge_cells(start_row=tr, end_row=tr, start_column=1, end_column=9)
    for c in range(2, 10):
        ws.cell(tr, c).border = A.border_all
    # Suma Valor Unit (J), IVA (K), Total c/IVA (L)
    for c in (10, 11, 12):
        L = get_column_letter(c)
        cc = ws.cell(tr, c, f"=SUM({L}{d0}:{L}{d1})")
        cc.fill = A.fill(A.TEAL); cc.font = A.font(11 if c == 12 else 10, True, A.WHITE)
        cc.alignment = A.center; cc.border = A.border_all; cc.number_format = '#,##0'
    ws.cell(tr, 13).border = A.border_all

    # ── Nota al pie ──
    ws.cell(tr + 2, 1, NOTA_DISC).font = A.font(9, False, "7A8094")
    ws.merge_cells(start_row=tr+2, end_row=tr+2,
                   start_column=1, end_column=NC)

    ws.sheet_properties.tabColor = '5B6BFF'

    print(f"📋 Listado Equipos OK: {N} seriales · filas {d0}-{d1}")


# ══════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv-det", required=True,
                    help="CSV granular por serial (resultado del SQL listado). "
                         "Ambas hojas se construyen desde este archivo.")
    ap.add_argument("--csv-agg", help="(opcional, ya no se usa: la Hoja 1 se agrega desde --csv-det)")
    ap.add_argument("--out", default=str(ROOT / "Calibraciones.xlsx"))
    args = ap.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    build_hoja1(wb, args.csv_det)   # agrega desde el detalle → cuadra con Hoja 2
    build_hoja2(wb, args.csv_det)

    wb.save(args.out)
    print(f"\n✅ Generado: {args.out}")
    print("   Hoja 1: 📅 Calibraciones (resumen mensual + tabla por tipo + detalle agregado)")
    print("   Hoja 2: 📋 Listado Equipos (726 seriales, precio unitario, vencimiento exacto)")


if __name__ == "__main__":
    main()
