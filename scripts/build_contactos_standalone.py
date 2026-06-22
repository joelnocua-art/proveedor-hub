#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_contactos_standalone.py — Genera SOLO la hoja "📞 Contactos"
en un archivo .xlsx propio, leyendo providers.json.

Mejoras vs. la versión embebida en build_excel.py:
  · Franja de KPIs con fórmulas vivas (total, con web, con correo, con tel.)
  · Columnas extra: Categoría, Dirección, NIT/RUT
  · Hipervínculos en Correo (mailto:) y Página Web
  · Columna "Completitud" = COUNTA de los 4 campos de contacto (fórmula viva)
  · Formato condicional: Estado ACTIVO verde · Completitud escala de color

Uso:
  python3 scripts/build_contactos_standalone.py
  python3 scripts/build_contactos_standalone.py --out MisContactos.xlsx
"""
import sys, json, argparse
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(Path(__file__).parent))

import add_sheet as A
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

LINK = "1A6FB8"  # azul para hipervínculos


def build(wb, providers):
    ws = wb.create_sheet("📞 Contactos")
    ws.sheet_view.showGridLines = False

    # (encabezado, ancho, clave en providers.json)
    COLS = [
        ("Empresa", 30, "empresa"),
        ("Categoría", 13, "categoria"),
        ("Estado", 10, "estado"),
        ("Representante", 22, "representante"),
        ("Celular / Tel.", 15, "celular_telefono"),
        ("Correo", 30, "correo_electronico"),
        ("Página Web", 30, "pagina_web"),
        ("Ciudad", 13, "ciudad"),
        ("Dirección", 26, "direccion"),
        ("NIT / RUT", 14, "rut"),
        ("Servicios / Productos", 40, "servicios_productos"),
        ("Completitud", 12, None),  # fórmula
    ]
    NC = len(COLS)
    COMPL = NC  # última columna

    A.banner(ws, "  📞 Contactos por Proveedor",
             "  Para responder '¿cómo contacto a X?' · correo y web son clickeables · Completitud = campos de contacto llenos /4",
             NC)

    # ── Filas de datos ──
    provs = sorted(providers, key=lambda p: (p.get('empresa') or '').lower())
    provs = [p for p in provs if (p.get('empresa') or '').upper() != 'PRUEBA']

    hdr = 4
    d0 = hdr + 1
    d1 = d0 + len(provs) - 1

    # ── Franja KPI (fila 3) con fórmulas vivas sobre el detalle ──
    # D=Representante E=Celular F=Correo G=Web
    kpis = [
        ("Proveedores", f"=COUNTA(A{d0}:A{d1})"),
        ("Con teléfono", f"=COUNTA(E{d0}:E{d1})"),
        ("Con correo",   f"=COUNTA(F{d0}:F{d1})"),
        ("Con web",      f"=COUNTA(G{d0}:G{d1})"),
    ]
    col = 1
    span = 3  # cada KPI ocupa 3 columnas (label arriba, valor abajo no — lo ponemos lado a lado)
    for label, formula in kpis:
        lc = ws.cell(3, col, label)
        lc.fill = A.fill(A.NAVY); lc.font = A.font(9, True, A.WHITE)
        lc.alignment = A.center; lc.border = A.border_all
        vc = ws.cell(3, col + 1, formula)
        vc.fill = A.fill(A.TEAL); vc.font = A.font(12, True, A.WHITE)
        vc.alignment = A.center; vc.border = A.border_all
        vc.number_format = '#,##0'
        col += span
    # rellenar resto de la fila 3
    for c in range(col, NC + 1):
        cc = ws.cell(3, c); cc.fill = A.fill(A.GREY_LT); cc.border = A.border_all
    ws.row_dimensions[3].height = 22

    # ── Header ──
    for i, (name, w, _) in enumerate(COLS, 1):
        ws.cell(hdr, i, name)
        ws.column_dimensions[get_column_letter(i)].width = w
    A.style_header(ws, NC, hdr)

    # ── Datos ──
    rr = d0
    for p in provs:
        for ci, (_, _, key) in enumerate(COLS, 1):
            if ci == COMPL:
                # Completitud = nº de campos de contacto llenos (Rep, Tel, Correo, Web)
                f = f'=COUNTA(D{rr}:G{rr})&" / 4"'
                cc = ws.cell(rr, ci, f)
                cc.alignment = A.center
            else:
                v = p.get(key) or None
                cc = ws.cell(rr, ci, v)
                cc.alignment = A.left if ci in (1, 4, 6, 7, 9, 11) else A.center
            cc.border = A.border_all
            cc.font = A.font(10)
        # Hipervínculos
        correo = p.get('correo_electronico')
        if correo:
            c = ws.cell(rr, 6)
            c.hyperlink = f"mailto:{correo}"
            c.font = A.font(10, False, LINK)
        web = p.get('pagina_web')
        if web:
            c = ws.cell(rr, 7)
            c.hyperlink = web if web.startswith('http') else f"https://{web}"
            c.font = A.font(10, False, LINK)
        rr += 1

    # ── Tabla ──
    tab = Table(displayName="Contactos", ref=f"A{hdr}:{get_column_letter(NC)}{d1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = f"A{d0}"

    # ── Formato condicional ──
    # Estado ACTIVO en verde
    ws.conditional_formatting.add(
        f"C{d0}:C{d1}",
        CellIsRule(operator='equal', formula=['"ACTIVO"'],
                   fill=A.fill(A.GREEN_LT), font=A.font(10, True, A.GREEN)))
    # Celdas de contacto vacías → resaltar en rojo claro (Rep, Tel, Correo, Web)
    for colL in ("E", "F", "G"):
        ws.conditional_formatting.add(
            f"{colL}{d0}:{colL}{d1}",
            CellIsRule(operator='equal', formula=['""'],
                       fill=A.fill(A.RED_LT)))

    # Nota al pie
    nota = ("Verde = proveedor ACTIVO · Celdas rojas = dato de contacto faltante (por completar) · "
            "Correo y Web son clickeables · Datos enriquecidos vía búsqueda web jun-2026")
    ws.cell(d1 + 2, 1, nota).font = A.font(9, False, "7A8094")
    ws.merge_cells(start_row=d1 + 2, end_row=d1 + 2, start_column=1, end_column=NC)

    ws.sheet_properties.tabColor = A.TAB_COLORS.get("📞 Contactos", A.GREEN)

    # Stats para consola
    con_web = sum(1 for p in provs if p.get('pagina_web'))
    con_cor = sum(1 for p in provs if p.get('correo_electronico'))
    con_tel = sum(1 for p in provs if p.get('celular_telefono'))
    print(f"📞 Contactos OK: {len(provs)} proveedores · web {con_web} · correo {con_cor} · tel {con_tel} · filas {d0}-{d1}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(ROOT / "Contactos.xlsx"))
    args = ap.parse_args()

    providers = json.load(open(ROOT / "providers.json"))

    wb = Workbook()
    wb.remove(wb.active)
    build(wb, providers)
    wb.save(args.out)
    print(f"\n✅ Generado: {args.out}")


if __name__ == "__main__":
    main()
