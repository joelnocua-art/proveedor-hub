# -*- coding: utf-8 -*-
"""
build_excel.py  —  BIA Energy · KB Proveedor Hub
Genera el Excel COMPLETO de 8 hojas con certificaciones ya inyectadas.

Uso directo:
    python3 scripts/build_excel.py
Uso desde metabase_sync.py:
    python3 scripts/build_excel.py --data /path/to/kb_data.json --out KB_BIA_Proveedores.xlsx
"""
import argparse, json, sys
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

ROOT = Path(__file__).parent.parent

_ap = argparse.ArgumentParser(add_help=False)
_ap.add_argument("--data", default=None)
_ap.add_argument("--out",  default=None)
_args, _ = _ap.parse_known_args()

_data_path = _args.data or (ROOT / "scripts" / "kb_data_cache.json")
if not Path(_data_path).exists():
    _data_path = "/tmp/kb_data.json"

D = json.load(open(_data_path))
rows = D['rows']; inv_by_sku = D['inv_by_sku']
sku_type = D['sku_type']; sku_name = D['sku_name']
lead_times = D['lead_times']; offers_by_sku = D['offers_by_sku']

# ════════════════════════════════════════════════════════════════
# CERTIFICACIONES INVESTIGADAS (por proveedor)
# Fuente: investigación web jun-2026 (CIDET, ONAC, SICERCO, sitios oficiales)
# ════════════════════════════════════════════════════════════════
PROV_CERT = {
    "SELDA":     {"creg":"Sí", "retie":"Sí", "homol":"Sí",
                  "nota":"Certificado CIDET (ONAC 09-CPR-004), cumple RETIE. Homologado por operadores"},
    "PROELCO":   {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Distribuidor con catálogo público. Solicitar certificado CREG/CIDET"},
    "EMSI TECH": {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Importador EMSI Ltd (China), marca MICROSTAR. Solicitar certificación"},
    "4S":        {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Sin info pública. Confirmar si distribuye medidores certificados"},
    "ACJ":       {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Sin info pública. Verificar línea de medición y certificados"},
    "ADLER":     {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Materiales/automatización. Solicitar certificados si aplica a medición"},
    "DISICO":    {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Fabricante TCs. Solicitar certificado de conformidad"},
    "LAUMAYER":  {"creg":"Por confirmar", "retie":"Por confirmar", "homol":"Por confirmar",
                  "nota":"Condensadores/tableros. Solicitar certificados si aplica"},
}
# Inyectar en rows (Catálogo Maestro)
for r in rows:
    p = r.get("Proveedor","").upper().strip()
    c = PROV_CERT.get(p)
    if c:
        r["CREG_038"] = c["creg"]
        r["RETIE_2024"] = c["retie"]
        r["Homologado"] = c["homol"]
        if not r.get("Detalle_Homologacion"):
            r["Detalle_Homologacion"] = c["nota"]

# ─── PALETA ──────────────────────────────────────────────────────
NAVY="1A2138";TEAL="00B8A0";TEAL_LT="D6F5F0";GOLD="E8A13A";GOLD_LT="FBEFD8"
RED="E24A5B";RED_LT="FBE3E6";GREEN="2EA56A";GREEN_LT="DDF3E8"
GREY_LT="F2F4F8";GREY_MD="E1E5EE";WHITE="FFFFFF";TXT="1A2138";YELLOW="FFF4CC"
def fill(c): return PatternFill("solid", fgColor=c)
def font(sz=11,b=False,color=TXT,name="Calibri"): return Font(name=name,size=sz,bold=b,color=color)
thin=Side(style="thin",color="C8CEDC")
border_all=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)
right=Alignment(horizontal="right",vertical="center")

wb=Workbook()
def style_header(ws,ncols,row=1,fillc=NAVY,txtc=WHITE,h=30):
    ws.row_dimensions[row].height=h
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c)
        cell.fill=fill(fillc);cell.font=font(11,True,txtc)
        cell.alignment=center;cell.border=border_all
def banner(ws,title,subtitle,ncols):
    ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=ncols)
    ws.merge_cells(start_row=2,end_row=2,start_column=1,end_column=ncols)
    t=ws.cell(1,1,title);t.fill=fill(NAVY);t.font=font(18,True,WHITE);t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    s=ws.cell(2,1,subtitle);s.fill=fill(TEAL);s.font=font(10,True,WHITE);s.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=34;ws.row_dimensions[2].height=20

TIPO_SIMPLE={'Medidor':'Medidor','Celda De Medidor':'Celda Medidor','Transformador De Corriente':'TC',
             'Transformador De Corriente BM':'TC','Transformador De Potencial':'TPS'}

# ════════════════════════════════════════════════════════════════
# HOJA 0 — INICIO
# ════════════════════════════════════════════════════════════════
ws=wb.active; ws.title="📘 Inicio"
ws.sheet_view.showGridLines=False
for col in range(1,8): ws.column_dimensions[get_column_letter(col)].width=[3,28,42,18,18,18,3][col-1]
banner(ws,"  BIA ENERGY · Knowledge Base de Proveedores","  Fuente de datos para el Copilot interno · Consolidado 15-jun-2026",7)
guide=[
    ("",""),
    ("¿QUÉ ES ESTE ARCHIVO?",""),
    ("","Toda la data de proveedores, precios, inventario y normativa en un solo lugar,"),
    ("","lista para alimentar el Copilot interno de BIA."),
    ("",""),
    ("HOJAS DEL ARCHIVO",""),
    ("🔎 Buscador","Escribe un SKU o proveedor y ve todo: precios, stock, normativa."),
    ("📋 Catálogo Maestro","Tabla completa ítem por ítem (985 filas). Filtrable por tipo, proveedor, país."),
    ("💰 Proveedor + Barato","Por cada equipo, quién lo vende más barato y el ahorro vs el más caro."),
    ("🚚 Lead Times","Tiempo de entrega promedio por proveedor (el más rápido y el más lento)."),
    ("📦 Inventario","Cuánto tenemos hoy por tipo: asignado, disponible, vencido y días que alcanza."),
    ("📝 Pendiente Jarri","Normativa CREG 038 / RETIE 2024 / Homologado. Ya pre-llenado lo investigado."),
    ("📈 Insights","Top equipos, alertas de stock crítico y resumen para decisiones rápidas."),
    ("",""),
    ("CÓMO LO USA EL COPILOT",""),
    ("","Preguntas como '¿proveedor más barato del medidor INHEMETER?' → Hoja 💰"),
    ("","'¿para cuántos días alcanza el stock de medidores?' → Hoja 📦"),
    ("","'¿cuánto se demora SELDA en entregar?' → Hoja 🚚"),
    ("","'¿este proveedor cumple CREG 038 / RETIE?' → Hoja 📝"),
    ("",""),
    ("ESTADO DE LOS DATOS",""),
    ("✅ Listo","Precios, inventario, lead times, país (922 precios reales, 13.527 equipos)."),
    ("✅ Investigado","SELDA confirmado CREG+RETIE+CIDET. Otros 'Por confirmar' (ver hoja 📝)."),
    ("⚠️ Validar","Jarri confirma 'Por confirmar' pidiendo certificado ONAC a cada proveedor."),
]
r=4
for a,b in guide:
    ca=ws.cell(r,2,a); cb=ws.cell(r,3,b)
    if a and not b:
        ca.font=font(12,True,TEAL); ws.merge_cells(start_row=r,end_row=r,start_column=2,end_column=6)
    else:
        ca.font=font(11,True,NAVY); cb.font=font(11,False,TXT)
        ws.merge_cells(start_row=r,end_row=r,start_column=3,end_column=6)
    ca.alignment=left; cb.alignment=left
    r+=1
ws.sheet_view.zoomScale=110

# ════════════════════════════════════════════════════════════════
# HOJA — CATÁLOGO MAESTRO
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("📋 Catálogo Maestro")
ws.sheet_view.showGridLines=False
COLS=[("Tipo",16),("Referencia / Modelo",46),("Proveedor",20),("País",13),
      ("Precio (COP)",15),("Lead Time",11),("Inv. Asignado",12),("Inv. Disponible",13),
      ("Inv. Vencido",11),("Pend. Cert.",11),("Días Inv.",10),("CREG 038",10),
      ("RETIE 2024",11),("Homologado",12),("Detalle Homologación",30)]
banner(ws,"  📋 Catálogo Maestro — Ítem por Ítem","  985 registros · Filtra con las flechas · Verde=barato Rojo=caro por SKU",len(COLS))
hr=3
for i,(n,w) in enumerate(COLS,1):
    ws.cell(hr,i,n); ws.column_dimensions[get_column_letter(i)].width=w
style_header(ws,len(COLS),row=hr,h=32)
KEYS=['Tipo','Referencia','Proveedor','Pais','Precio_COP','LeadTime_Dias',
      'Inv_Asignado','Inv_Disponible','Inv_Vencido','Pend_Certificado',
      'Dias_Inventario','CREG_038','RETIE_2024','Homologado','Detalle_Homologacion']
def sortkey(r):
    pr=r['Precio_COP'] if isinstance(r['Precio_COP'],(int,float)) else 9e18
    return (r['Tipo'],r['Referencia'],pr)
srows=sorted(rows,key=sortkey)
rr=hr+1
for r in srows:
    for ci,k in enumerate(KEYS,1):
        v=r.get(k,'')
        cell=ws.cell(rr,ci,v if v!='' else None)
        cell.border=border_all; cell.font=font(10)
        if k=='Precio_COP': cell.alignment=right; cell.number_format='#,##0'
        elif k in ('Referencia','Detalle_Homologacion'): cell.alignment=left
        else: cell.alignment=center
    rr+=1
last=rr-1
tab=Table(displayName="Catalogo",ref=f"A{hr}:{get_column_letter(len(COLS))}{last}")
tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True,showColumnStripes=False)
ws.add_table(tab)
ws.freeze_panes=f"C{hr+1}"
ws.conditional_formatting.add(f"E{hr+1}:E{last}",
    ColorScaleRule(start_type='min',start_color=GREEN_LT,end_type='max',end_color=RED_LT))
ws.conditional_formatting.add(f"I{hr+1}:I{last}",
    CellIsRule(operator='greaterThan',formula=['0'],fill=fill(RED_LT),font=font(10,True,RED)))
# Colorear CREG/RETIE/Homologado (L,M,N)
for col in ("L","M","N"):
    ws.conditional_formatting.add(f"{col}{hr+1}:{col}{last}",
        CellIsRule(operator='equal',formula=['"Sí"'],fill=fill(GREEN_LT),font=font(10,True,GREEN)))
    ws.conditional_formatting.add(f"{col}{hr+1}:{col}{last}",
        CellIsRule(operator='equal',formula=['"No"'],fill=fill(RED_LT),font=font(10,True,RED)))
CAT_LAST=last
print("Catálogo OK:",last-hr,"filas")

# ════════════════════════════════════════════════════════════════
# HOJA — PROVEEDOR + BARATO
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("💰 Proveedor + Barato")
ws.sheet_view.showGridLines=False
COLS=[("Tipo",16),("Referencia / Modelo",48),("🏆 Más Barato",20),("Precio Mín (COP)",16),
      ("Precio Máx (COP)",16),("Ahorro %",11),("# Proveedores",13),("Todos los precios",55)]
banner(ws,"  💰 Proveedor Más Barato por Equipo","  Para responder: '¿quién vende más barato el SKU X?' · Ordenado por mayor ahorro",len(COLS))
for i,(n,w) in enumerate(COLS,1):
    ws.cell(3,i,n); ws.column_dimensions[get_column_letter(i)].width=w
style_header(ws,len(COLS),3)
best=[]
for s_id,offs in offers_by_sku.items():
    valid=[o for o in offs if o['price']]
    if not valid: continue
    valid.sort(key=lambda x:x['price'])
    cheap=valid[0]; dear=valid[-1]
    ahorro=round((1-cheap['price']/dear['price'])*100,0) if dear['price'] else 0
    todos=" · ".join(f"{o['provider']}: ${int(o['price']):,}" for o in valid)
    best.append({'tipo':TIPO_SIMPLE.get(sku_type.get(s_id,''),sku_type.get(s_id,'Otros')),
                 'ref':sku_name.get(s_id,'?'),'best':cheap['provider'],'min':int(cheap['price']),
                 'max':int(dear['price']),'ahorro':ahorro,'n':len(valid),'todos':todos})
best.sort(key=lambda x:(-x['ahorro'],-x['max']))
rr=4
for b in best:
    ws.cell(rr,1,b['tipo']).alignment=center
    ws.cell(rr,2,b['ref']).alignment=left
    ws.cell(rr,3,b['best']).alignment=center
    c4=ws.cell(rr,4,b['min']); c4.alignment=right; c4.number_format='#,##0'
    c5=ws.cell(rr,5,b['max']); c5.alignment=right; c5.number_format='#,##0'
    c6=ws.cell(rr,6,b['ahorro']/100 if b['ahorro'] else 0); c6.alignment=center; c6.number_format='0%'
    ws.cell(rr,7,b['n']).alignment=center
    ws.cell(rr,8,b['todos']).alignment=left
    for ci in range(1,9):
        cc=ws.cell(rr,ci); cc.border=border_all; cc.font=font(10)
    ws.cell(rr,3).font=font(10,True,GREEN)
    rr+=1
last=rr-1
tab=Table(displayName="MasBarato",ref=f"A3:H{last}")
tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes="A4"
ws.conditional_formatting.add(f"F4:F{last}",
    ColorScaleRule(start_type='num',start_value=0,start_color=WHITE,end_type='num',end_value=0.7,end_color=GREEN))
print("Proveedor+barato OK:",last-3)

# ════════════════════════════════════════════════════════════════
# HOJA — LEAD TIMES
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("🚚 Lead Times")
ws.sheet_view.showGridLines=False
COLS=[("Proveedor",30),("Lead Time (días)",18),("Velocidad",22),("# SKUs que cotiza",18)]
banner(ws,"  🚚 Tiempo de Entrega por Proveedor","  Para responder: '¿cuánto se demora X en entregar?' · Menor = más rápido",len(COLS))
for i,(n,w) in enumerate(COLS,1):
    ws.cell(3,i,n); ws.column_dimensions[get_column_letter(i)].width=w
style_header(ws,len(COLS),3)
prov_sku_count=defaultdict(set)
for sid,offs in offers_by_sku.items():
    for o in offs:
        if o['price']: prov_sku_count[o['provider'].upper().strip()].add(sid)
clean_lt={k:v for k,v in lead_times.items() if 'PRUEB' not in k and 'TEST' not in k}
lt_sorted=sorted(clean_lt.items(),key=lambda x:x[1])
rr=4
for prov,days in lt_sorted:
    cnt=0
    for pk,sids in prov_sku_count.items():
        pc=pk.replace(' SAS','').replace(' S.A.S.','').replace(' INGENIERIA','').replace('-','').replace(' - ','').strip()
        kc=prov.replace(' SAS','').replace(' S.A.S.','').replace(' INGENIERIA','').replace('-','').replace(' - ','').strip()
        if pc==kc or pc in kc or kc in pc: cnt=max(cnt,len(sids))
    vel="🟢 Muy rápido" if days<=5 else ("🟡 Medio" if days<=15 else "🔴 Lento")
    ws.cell(rr,1,prov).alignment=left
    ws.cell(rr,2,days).alignment=center
    ws.cell(rr,3,vel).alignment=center
    ws.cell(rr,4,cnt if cnt else "—").alignment=center
    for ci in range(1,5):
        cc=ws.cell(rr,ci); cc.border=border_all; cc.font=font(11)
    rr+=1
last=rr-1
ws.conditional_formatting.add(f"B4:B{last}",
    ColorScaleRule(start_type='min',start_color=GREEN_LT,mid_type='percentile',mid_value=50,mid_color=GOLD_LT,end_type='max',end_color=RED_LT))
ws.freeze_panes="A4"
ws.cell(last+2,1,"Nota: Lead time = promedio de días de entrega histórico (tabla provider_responses).").font=font(9,False,"7A8094")
ws.merge_cells(start_row=last+2,end_row=last+2,start_column=1,end_column=4)
print("Lead times OK:",last-3)

# ════════════════════════════════════════════════════════════════
# HOJA — INVENTARIO
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("📦 Inventario")
ws.sheet_view.showGridLines=False
type_stats=defaultdict(lambda:{'asignado':0,'disponible':0,'vencido':0,'pend':0,'total':0})
for sid,iv in inv_by_sku.items():
    cat=sku_type.get(sid,'Otros'); t=TIPO_SIMPLE.get(cat,cat)
    s=type_stats[t]
    s['asignado']+=iv['asignado']; s['disponible']+=iv['disponible']
    s['vencido']+=iv['vencido']; s['pend']+=iv.get('pend_cert',0); s['total']+=iv['total']
banner(ws,"  📦 Inventario Actual — ¿Para cuántos días alcanza?","  Asignado / Disponible / Vencido por tipo · Días = Disponible ÷ Consumo diario",8)
ws.cell(3,1,"RESUMEN POR TIPO DE EQUIPO").font=font(13,True,NAVY)
ws.merge_cells(start_row=3,end_row=3,start_column=1,end_column=8)
COLS=[("Tipo de Equipo",22),("Total",10),("Asignado",11),("Disponible",12),("Vencido",10),
      ("Pend. Cert.",12),("Consumo/día",12),("Días Inventario",15)]
for i,(n,w) in enumerate(COLS,1):
    ws.cell(4,i,n); ws.column_dimensions[get_column_letter(i)].width=w
style_header(ws,8,4)
rr=5
order=['Medidor','TPS','TC','Celda Medidor','Banco Condensador','Condensador']
ordered=[t for t in order if t in type_stats]+[t for t in type_stats if t not in order]
for t in ordered:
    s=type_stats[t]
    cdp=round(s['asignado']/180,2) if s['asignado'] else 0
    dias=round(s['disponible']/cdp,0) if cdp else 0
    ws.cell(rr,1,t).alignment=left; ws.cell(rr,1).font=font(11,True)
    ws.cell(rr,2,s['total']).alignment=center
    ws.cell(rr,3,s['asignado']).alignment=center
    ws.cell(rr,4,s['disponible']).alignment=center
    ws.cell(rr,5,s['vencido']).alignment=center
    ws.cell(rr,6,s['pend']).alignment=center
    ws.cell(rr,7,cdp).alignment=center
    cd=ws.cell(rr,8,dias if dias else "—"); cd.alignment=center; cd.font=font(11,True)
    for ci in range(1,9):
        cc=ws.cell(rr,ci); cc.border=border_all
        if not cc.font.bold: cc.font=font(11)
    rr+=1
sumrow=rr-1
ws.conditional_formatting.add(f"H5:H{sumrow}",
    ColorScaleRule(start_type='num',start_value=0,start_color=RED,mid_type='num',mid_value=90,mid_color=GOLD,end_type='num',end_value=365,end_color=GREEN))
ws.conditional_formatting.add(f"E5:E{sumrow}",
    CellIsRule(operator='greaterThan',formula=['0'],fill=fill(RED_LT),font=font(11,True,RED)))
det_start=sumrow+3
ws.cell(det_start,1,"DETALLE POR SKU (con inventario)").font=font(13,True,NAVY)
ws.merge_cells(start_row=det_start,end_row=det_start,start_column=1,end_column=8)
hr=det_start+1
DCOLS=[("Tipo",16),("Referencia / Modelo",50),("Asignado",11),("Disponible",12),
       ("Vencido",10),("Pend. Cert.",12),("Consumo/día",12),("Días Inv.",11)]
for i,(n,w) in enumerate(DCOLS,1):
    ws.cell(hr,i,n)
style_header(ws,8,hr)
rr=hr+1
det=[]
for sid,iv in inv_by_sku.items():
    if iv['total']==0: continue
    cat=sku_type.get(sid,'Otros'); t=TIPO_SIMPLE.get(cat,cat)
    cdp=round(iv['asignado']/180,3) if iv['asignado'] else 0
    dias=round(iv['disponible']/cdp,0) if cdp else ''
    det.append((t,sku_name.get(sid,'?'),iv['asignado'],iv['disponible'],iv['vencido'],iv.get('pend_cert',0),cdp,dias))
det.sort(key=lambda x:-(x[2]+x[3]))
for d in det:
    for ci,v in enumerate(d,1):
        cc=ws.cell(rr,ci,v if v!='' else None); cc.border=border_all; cc.font=font(10)
        cc.alignment=left if ci==2 else center
    rr+=1
dlast=rr-1
tab=Table(displayName="InvDetalle",ref=f"A{hr}:H{dlast}")
tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes=f"A{hr+1}"
print("Inventario OK:",dlast-hr,"SKUs detalle")

# ════════════════════════════════════════════════════════════════
# HOJA — PENDIENTE JARRI (con resumen de proveedores PRE-LLENADO)
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("📝 Pendiente Jarri")
ws.sheet_view.showGridLines=False
COLS=[("Tipo",16),("Referencia / Modelo",52),("CREG 038",12),("RETIE 2024",12),
      ("Homologado",12),("Detalle / Notas",46)]
for i,(n,w) in enumerate(COLS,1):
    ws.column_dimensions[get_column_letter(i)].width=w
banner(ws,"  📝 Normativa — CREG 038 / RETIE 2024 / Homologado","  Pre-llenado con investigación web (jun-2026). Jarri valida 'Por confirmar' con cada proveedor",len(COLS))

# --- Bloque resumen por PROVEEDOR (lo investigado) ---
ws.cell(3,1,"RESUMEN POR PROVEEDOR (investigado)").font=font(13,True,NAVY)
ws.merge_cells(start_row=3,end_row=3,start_column=1,end_column=6)
hdr_prov=["Proveedor","Línea / Producto","CREG 038","RETIE 2024","Homologado","Fuente / Nota"]
for i,h in enumerate(hdr_prov,1):
    c=ws.cell(4,i,h); c.fill=fill(NAVY); c.font=font(11,True,WHITE); c.alignment=center; c.border=border_all
ws.row_dimensions[4].height=26
prov_lines={"SELDA":"Medidores, TC, TPS","PROELCO":"Medidores y transformadores",
            "EMSI TECH":"Medidores (MICROSTAR)","4S":"Ingeniería eléctrica",
            "ACJ":"Por verificar","ADLER":"Materiales/automatización",
            "DISICO":"Transformadores de corriente","LAUMAYER":"Condensadores/tableros"}
prov_order=["SELDA","PROELCO","EMSI TECH","4S","ACJ","ADLER","DISICO","LAUMAYER"]
rr=5
def cert_cell(cell,val):
    cell.alignment=center; cell.border=border_all
    if val=="Sí": cell.font=font(10,True,GREEN); cell.fill=fill(GREEN_LT)
    elif val=="No": cell.font=font(10,True,RED); cell.fill=fill(RED_LT)
    else: cell.font=font(10,True,GOLD); cell.fill=fill(GOLD_LT)
for p in prov_order:
    c=PROV_CERT[p]
    ws.cell(rr,1,p).alignment=left; ws.cell(rr,1).font=font(10,True); ws.cell(rr,1).border=border_all
    ws.cell(rr,2,prov_lines.get(p,"")).alignment=left; ws.cell(rr,2).font=font(10); ws.cell(rr,2).border=border_all
    cert_cell(ws.cell(rr,3,c["creg"]),c["creg"])
    cert_cell(ws.cell(rr,4,c["retie"]),c["retie"])
    cert_cell(ws.cell(rr,5,c["homol"]),c["homol"])
    ws.cell(rr,6,c["nota"]).alignment=left; ws.cell(rr,6).font=font(9); ws.cell(rr,6).border=border_all
    rr+=1
prov_last=rr-1

# --- Detalle por SKU activo de medición (pre-llenado por proveedor más barato) ---
det_start=prov_last+2
ws.cell(det_start,1,"DETALLE POR SKU ACTIVO (medición) — pre-llenado; ajusta si es necesario").font=font(13,True,NAVY)
ws.merge_cells(start_row=det_start,end_row=det_start,start_column=1,end_column=6)
hr=det_start+1
for i,(n,w) in enumerate(COLS,1):
    ws.cell(hr,i,n)
style_header(ws,len(COLS),hr)
RELEVANT={'Medidor','TC','TPS','Celda Medidor','Banco Condensador','Condensador'}
seen=set(); jarri=[]
for sid,offs in offers_by_sku.items():
    valid=[o for o in offs if o['price']]
    if not valid: continue
    cat=sku_type.get(sid,'Otros'); t=TIPO_SIMPLE.get(cat,cat)
    if t not in RELEVANT: continue
    ref=sku_name.get(sid,'?')
    if ref in seen: continue
    seen.add(ref)
    valid.sort(key=lambda x:x['price'])
    cheapest=valid[0]['provider'].upper().strip()
    jarri.append((t,ref,cheapest))
torder={'Medidor':0,'TPS':1,'TC':2,'Celda Medidor':3}
jarri.sort(key=lambda x:(torder.get(x[0],9),x[1]))
rr=hr+1
for t,ref,prov in jarri:
    c=PROV_CERT.get(prov,{"creg":"Por confirmar","retie":"Por confirmar","homol":"Por confirmar","nota":""})
    ws.cell(rr,1,t).alignment=center; ws.cell(rr,1).font=font(10,True)
    ws.cell(rr,2,ref).alignment=left; ws.cell(rr,2).font=font(10)
    ws.cell(rr,3,c["creg"]).alignment=center; ws.cell(rr,3).font=font(10)
    ws.cell(rr,4,c["retie"]).alignment=center; ws.cell(rr,4).font=font(10)
    ws.cell(rr,5,c["homol"]).alignment=center; ws.cell(rr,5).font=font(10)
    nota=f"Más barato: {prov}. {c.get('nota','')}"
    ws.cell(rr,6,nota).alignment=left; ws.cell(rr,6).font=font(9)
    for ci in range(1,7): ws.cell(rr,ci).border=border_all
    rr+=1
jlast=rr-1
dv=DataValidation(type="list",formula1='"Sí,No,Por confirmar"',allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"C{hr+1}:E{jlast}")
tab=Table(displayName="Jarri",ref=f"A{hr}:F{jlast}")
tab.tableStyleInfo=TableStyleInfo(name="TableStyleLight16",showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes=f"A{hr+1}"
ws.conditional_formatting.add(f"C{hr+1}:E{jlast}",CellIsRule(operator='equal',formula=['"Sí"'],fill=fill(GREEN_LT),font=font(10,True,GREEN)))
ws.conditional_formatting.add(f"C{hr+1}:E{jlast}",CellIsRule(operator='equal',formula=['"No"'],fill=fill(RED_LT),font=font(10,True,RED)))
ws.conditional_formatting.add(f"C{hr+1}:E{jlast}",CellIsRule(operator='equal',formula=['"Por confirmar"'],fill=fill(GOLD_LT),font=font(10,True,GOLD)))
print("Pendiente Jarri OK:",jlast-hr,"SKUs +",prov_last-4,"proveedores")

# ════════════════════════════════════════════════════════════════
# HOJA — INSIGHTS
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("📈 Insights")
ws.sheet_view.showGridLines=False
for c in range(1,9): ws.column_dimensions[get_column_letter(c)].width=[3,30,14,14,14,14,14,3][c-1]
banner(ws,"  📈 Insights & Alertas","  Resumen ejecutivo para decisiones rápidas",8)
ins=defaultdict(int)
for sid,iv in inv_by_sku.items(): ins[sid]=iv['asignado']
def section(ws,r,title):
    ws.cell(r,2,title).font=font(13,True,TEAL)
    ws.merge_cells(start_row=r,end_row=r,start_column=2,end_column=7)
    return r+1
r=4
r=section(ws,r,"🏆 Top 8 equipos más instalados (mayor demanda)")
ws.cell(r,2,"Equipo").font=font(10,True,WHITE);ws.cell(r,2).fill=fill(NAVY);ws.cell(r,2).border=border_all;ws.cell(r,2).alignment=left
ws.cell(r,3,"Instalados").font=font(10,True,WHITE);ws.cell(r,3).fill=fill(NAVY);ws.cell(r,3).border=border_all;ws.cell(r,3).alignment=center
ws.merge_cells(start_row=r,end_row=r,start_column=3,end_column=4)
r+=1
for sid,n in sorted(ins.items(),key=lambda x:-x[1])[:8]:
    ws.cell(r,2,sku_name.get(sid,'?')[:50]).alignment=left;ws.cell(r,2).font=font(10);ws.cell(r,2).border=border_all
    cc=ws.cell(r,3,n);cc.alignment=center;cc.font=font(10,True);cc.border=border_all
    ws.merge_cells(start_row=r,end_row=r,start_column=3,end_column=4)
    ws.cell(r,4).border=border_all
    r+=1
r+=1
r=section(ws,r,"⚠️ Alertas de stock crítico (poco disponible vs instalado)")
alerts=[]
for sid,iv in inv_by_sku.items():
    if iv['asignado']>=50 and iv['disponible']<iv['asignado']*0.1:
        alerts.append((sku_name.get(sid,'?'),iv['asignado'],iv['disponible']))
alerts.sort(key=lambda x:-x[1])
ws.cell(r,2,"Equipo").font=font(10,True,WHITE);ws.cell(r,2).fill=fill(RED);ws.cell(r,2).border=border_all;ws.cell(r,2).alignment=left
ws.cell(r,3,"Instalado").font=font(10,True,WHITE);ws.cell(r,3).fill=fill(RED);ws.cell(r,3).border=border_all;ws.cell(r,3).alignment=center
ws.cell(r,4,"Disponible").font=font(10,True,WHITE);ws.cell(r,4).fill=fill(RED);ws.cell(r,4).border=border_all;ws.cell(r,4).alignment=center
r+=1
for ref,a,d in alerts[:8]:
    ws.cell(r,2,ref[:50]).alignment=left;ws.cell(r,2).font=font(10);ws.cell(r,2).border=border_all
    ws.cell(r,3,a).alignment=center;ws.cell(r,3).font=font(10);ws.cell(r,3).border=border_all
    cc=ws.cell(r,4,d);cc.alignment=center;cc.font=font(10,True,RED);cc.border=border_all
    r+=1
r+=1
r=section(ws,r,"📊 Totales del inventario (13.527 equipos)")
totals=[("Total equipos",sum(iv['total'] for iv in inv_by_sku.values())),
        ("Instalados / Asignados",sum(iv['asignado'] for iv in inv_by_sku.values())),
        ("Disponibles en bodega",sum(iv['disponible'] for iv in inv_by_sku.values())),
        ("Averiados / Vencidos",sum(iv['vencido'] for iv in inv_by_sku.values())),
        ("Pendiente certificados",sum(iv.get('pend_cert',0) for iv in inv_by_sku.values()))]
for name,val in totals:
    ws.cell(r,2,name).font=font(11);ws.cell(r,2).alignment=left;ws.cell(r,2).border=border_all
    cc=ws.cell(r,3,val);cc.font=font(11,True,NAVY);cc.alignment=center;cc.border=border_all;cc.number_format='#,##0'
    ws.merge_cells(start_row=r,end_row=r,start_column=3,end_column=4)
    ws.cell(r,4).border=border_all
    r+=1
print("Insights OK")

# ════════════════════════════════════════════════════════════════
# HOJA — BUSCADOR (al final, usando referencia al Catálogo)
# ════════════════════════════════════════════════════════════════
ws=wb.create_sheet("🔎 Buscador")
ws.sheet_view.showGridLines=False
CAT="'📋 Catálogo Maestro'"
HEADERS=["Tipo","Referencia / Modelo","Proveedor","País","Precio (COP)","Lead Time",
         "Inv. Asig.","Inv. Disp.","Inv. Venc.","Pend.Cert","Días Inv.","CREG 038","RETIE 2024","Homologado","Detalle Homolog."]
WIDTHS=[15,44,18,12,14,10,9,9,9,9,9,9,10,11,24]
for i,w in enumerate(WIDTHS,1): ws.column_dimensions[get_column_letter(i)].width=w
banner(ws,"  🔎 Buscador Rápido","  Escribe abajo un equipo o proveedor (ej: INHEMETER, TC 300, SELDA) y Enter",len(HEADERS))
ws.cell(4,1,"BUSCAR:").font=font(12,True,NAVY); ws.cell(4,1).alignment=Alignment(horizontal="right",vertical="center")
box=ws.cell(4,2,"SELDA")
box.fill=fill(YELLOW); box.font=font(13,True,TXT); box.alignment=Alignment(horizontal="left",vertical="center",indent=1)
box.border=Border(left=Side(style="medium",color=GOLD),right=Side(style="medium",color=GOLD),top=Side(style="medium",color=GOLD),bottom=Side(style="medium",color=GOLD))
ws.merge_cells("B4:E4")
ws.cell(4,6,"← cambia esta celda").font=font(9,False,"7A8094"); ws.cell(4,6).alignment=left
ws.merge_cells("F4:H4")
ws.row_dimensions[4].height=26
for i,h in enumerate(HEADERS,1):
    c=ws.cell(6,i,h); c.fill=fill(NAVY); c.font=font(10,True,WHITE); c.alignment=center; c.border=border_all
ws.row_dimensions[6].height=28
hr_cat=3  # header del catálogo en fila 3, datos desde 4
rng=lambda col: f"{CAT}!{col}4:{col}{CAT_LAST}"
formula=(f'=IFERROR(FILTER({CAT}!A4:O{CAT_LAST}, '
         f'ISNUMBER(SEARCH($B$4,{rng("B")}))+ISNUMBER(SEARCH($B$4,{rng("C")}))+ISNUMBER(SEARCH($B$4,{rng("A")}))), '
         f'"Sin resultados — revisa la palabra")')
ws.cell(7,1,formula)
ws.freeze_panes="A7"
for r in range(7,61):
    for c in range(1,len(HEADERS)+1):
        cell=ws.cell(r,c); cell.border=border_all; cell.font=font(10)
        cell.alignment=left if c in (2,15) else center
        if c==5: cell.number_format='#,##0'
ws.cell(62,1,"💡 Tip: el buscador trae coincidencias por Referencia, Proveedor o Tipo. Funciona en Google Sheets y Excel 365.").font=font(9,False,"7A8094")
ws.merge_cells("A62:O62")
print("Buscador OK")

# ════════════════════════════════════════════════════════════════
# REORDENAR + COLORES DE PESTAÑA
# ════════════════════════════════════════════════════════════════
order=["📘 Inicio","🔎 Buscador","📋 Catálogo Maestro","💰 Proveedor + Barato",
       "🚚 Lead Times","📦 Inventario","📝 Pendiente Jarri","📈 Insights"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
colors={"📘 Inicio":NAVY,"🔎 Buscador":GOLD,"📋 Catálogo Maestro":TEAL,
        "💰 Proveedor + Barato":GREEN,"🚚 Lead Times":"5B6BFF","📦 Inventario":"8855CC",
        "📝 Pendiente Jarri":RED,"📈 Insights":NAVY}
for s in wb._sheets:
    if s.title in colors: s.sheet_properties.tabColor=colors[s.title]

OUT = _args.out or str(ROOT / "KB_BIA_Proveedores.xlsx")
wb.save(OUT)
print("✓ Excel COMPLETO guardado:", OUT)
print("Hojas:", [s.title for s in wb._sheets])
